from tkinter import *
from tkinter import messagebox, ttk
import pandas as pd
from datetime import datetime
import os, shutil, threading

# ── Webcam imports (graceful fallback if not installed) ──
try:
    import cv2
    from PIL import Image, ImageTk
    WEBCAM_AVAILABLE = True
except ImportError:
    WEBCAM_AVAILABLE = False

# ═══════════════════════════════════════════
#  FILES & PATHS — dynamic, admin-configurable
# ═══════════════════════════════════════════
DATE_FORMAT = "%d-%m-%Y"
TIME_FORMAT = "%H:%M:%S"
OUT_EMPTY   = ["", "nan", "NaN", "None"]

import sys as _sys
if getattr(_sys, "frozen", False):
    _SCRIPT_DIR = os.path.dirname(os.path.abspath(_sys.executable))
else:
    _SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

SETTINGS_FILE = os.path.join(_SCRIPT_DIR, "GST_Settings.txt")
VIP_CONFIG_FILE = os.path.join(_SCRIPT_DIR, "GST_VIP_Config.txt")

def load_vip_enabled():
    """Return True if VIP system is ON, False if OFF. Default ON."""
    try:
        if os.path.exists(VIP_CONFIG_FILE):
            with open(VIP_CONFIG_FILE, "r", encoding="utf-8") as f:
                val = f.read().strip()
            return val != "OFF"
    except: pass
    return True   # default ON

def save_vip_enabled(enabled: bool):
    """Save VIP ON/OFF setting."""
    try:
        with open(VIP_CONFIG_FILE, "w", encoding="utf-8") as f:
            f.write("ON" if enabled else "OFF")
        return True
    except:
        return False

# ── Version & Auto-Update ──────────────────────────────────────
CURRENT_VERSION = "v15"
_GH_RAW      = ("https://raw.githubusercontent.com/"
                 "udbhavkumar05-stack/gst-vms-updates/main/")
VERSION_URL  = _GH_RAW + "version.txt"
# Two download URLs — one for .exe, one for .py
DOWNLOAD_URL_PY  = _GH_RAW + "gst_visitor_system_latest.py"
DOWNLOAD_URL_EXE = _GH_RAW + "VMS.exe"

def _get_current_file():
    """Return path to currently running file — works for both .exe and .py"""
    if getattr(_sys, "frozen", False):
        # Running as .exe (PyInstaller)
        return os.path.abspath(_sys.executable)
    else:
        # Running as .py script
        return os.path.abspath(__file__)

def _is_exe():
    return getattr(_sys, "frozen", False)

def _do_update_check(parent_win, manual=False):
    """
    manual=False -> silent when up to date (auto startup check)
    manual=True  -> always shows result (sidebar button click)
    """
    try:
        import urllib.request as _ur
        with _ur.urlopen(VERSION_URL, timeout=8) as r:
            raw = r.read().decode("utf-8").strip()
        lines     = raw.splitlines()
        new_ver   = lines[0].strip()
        changelog = "\n".join(lines[1:]).strip() if len(lines) > 1 else ""

        if new_ver == CURRENT_VERSION:
            if manual:
                # Check internet status
                try:
                    import urllib.request as _ur2
                    _ur2.urlopen("http://www.google.com", timeout=3)
                    net_status = "🌐 Internet: Online ✅"
                except:
                    net_status = "🌐 Internet: Offline ⚠️"
                parent_win.after(0, lambda: messagebox.showinfo(
                    "✅  Up to Date",
                    f"GST VMS is up to date.\n\n"
                    f"Current version : {CURRENT_VERSION}\n\n"
                    f"No update available right now.\n\n"
                    f"{net_status}",
                    parent=parent_win))
            return

        parent_win.after(0, lambda: _show_update_popup(
            parent_win, new_ver, changelog))

    except Exception as _ue:
        if manual:
            parent_win.after(0, lambda e=str(_ue): messagebox.showerror(
                "\u274c  Check Failed",
                f"Could not check for updates.\n\n"
                f"Reason: {e}\n\n"
                f"Check your internet connection and try again.",
                parent=parent_win))
        else:
            print(f"Auto update check: {_ue}")


def _show_update_popup(parent_win, new_ver, changelog):
    """Dead simple popup — linear pack only, no place, no side=BOTTOM."""
    top = Toplevel(parent_win)
    top.title("Update Available")
    top.configure(bg="#FFFFFF")
    top.resizable(True, True)
    top.transient(parent_win)
    top.grab_set()
    top.update_idletasks()
    sw = top.winfo_screenwidth()
    sh = top.winfo_screenheight()
    top.geometry(f"460x440+{(sw-460)//2}+{(sh-440)//2}")

    # 1. Header
    hdr = Frame(top, bg="#1B2E4B")
    hdr.pack(fill=X)
    Label(hdr, text="Update Available",
          font=("Segoe UI",15,"bold"),
          bg="#1B2E4B", fg="#C8A84B",
          pady=16).pack()

    # 2. Version box
    vbox = Frame(top, bg="#F0FDF4", padx=30, pady=20)
    vbox.pack(fill=X, padx=20, pady=20)
    Label(vbox, text="Current version",
          font=("Segoe UI",9), bg="#F0FDF4", fg="#64748B").pack(anchor=W)
    Label(vbox, text=CURRENT_VERSION,
          font=("Segoe UI",22,"bold"),
          bg="#F0FDF4", fg="#1E293B").pack(anchor=W)
    Label(vbox, text=" ",
          bg="#F0FDF4", fg="#F0FDF4").pack()
    Label(vbox, text="New version available",
          font=("Segoe UI",9), bg="#F0FDF4", fg="#64748B").pack(anchor=W)
    Label(vbox, text=new_ver,
          font=("Segoe UI",22,"bold"),
          bg="#F0FDF4", fg="#16A34A").pack(anchor=W)

    # 3. Safety note
    Label(top,
          text="Your Excel data and photos will NOT be changed.",
          font=("Segoe UI",9,"bold"),
          bg="#FFFFFF", fg="#16A34A").pack(pady=(0,10))

    # 4. Progress
    prog_var = StringVar(value="")
    Label(top, textvariable=prog_var,
          font=("Segoe UI",9,"italic"),
          bg="#FFFFFF", fg="#2563EB").pack()

    # 5. Buttons
    brow = Frame(top, bg="#FFFFFF")
    brow.pack(pady=16)

    def _download():
        upd.config(state=DISABLED, text="Downloading...")
        skp.config(state=DISABLED)
        prog_var.set("Downloading from GitHub...")
        top.update()
        def _dl():
            tmp = None
            try:
                import urllib.request as _ur, shutil as _sh
                dest        = _get_current_file()
                running_exe = _is_exe()
                dl_url      = DOWNLOAD_URL_EXE if running_exe else DOWNLOAD_URL_PY
                tmp         = dest + ".update_tmp"
                _ur.urlretrieve(dl_url, tmp)
                if os.path.getsize(tmp) < 5000:
                    raise Exception("Incomplete download")
                _sh.copy2(dest, dest + ".backup")
                if running_exe:
                    old = dest + ".old"
                    if os.path.exists(old): os.remove(old)
                    os.rename(dest, old)
                    os.rename(tmp, dest)
                else:
                    _sh.move(tmp, dest)
                top.after(0, _ok)
            except Exception as ex:
                try:
                    if tmp and os.path.exists(tmp): os.remove(tmp)
                except: pass
                top.after(0, lambda e=str(ex): _fail(e))
        def _ok():
            prog_var.set("Download complete!")
            messagebox.showinfo("Update Complete",
                f"Updated to {new_ver}!\n\n"
                f"Please CLOSE and RESTART the software.\n\n"
                f"Your data is safe.", parent=top)
            top.destroy()
        def _fail(e):
            prog_var.set(f"Failed: {e[:60]}")
            upd.config(state=NORMAL, text="Retry")
            skp.config(state=NORMAL)
            messagebox.showerror("Failed",
                f"Could not download.\nReason: {e}\n\nCheck internet.",
                parent=top)
        threading.Thread(target=_dl, daemon=True).start()

    upd = Button(brow, text="Update Now",
                 font=("Segoe UI",12,"bold"),
                 bg="#16A34A", fg="#FFFFFF",
                 relief=FLAT, cursor="hand2",
                 padx=28, pady=12,
                 activebackground="#15803D",
                 command=_download)
    upd.pack(side=LEFT, padx=(0,14))

    skp = Button(brow, text="Later",
                 font=("Segoe UI",11),
                 bg="#E2E8F0", fg="#475569",
                 relief=FLAT, cursor="hand2",
                 padx=20, pady=12,
                 activebackground="#CBD5E1",
                 command=top.destroy)
    skp.pack(side=LEFT)

def check_for_update_async(parent_win, manual=False):
    """Background update check — zero UI freeze."""
    threading.Thread(
        target=_do_update_check,
        args=(parent_win,),
        kwargs={"manual": manual},
        daemon=True).start()


def load_data_path():
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
                for line in f:
                    if line.startswith("DATA_PATH="):
                        p = line.strip().replace("DATA_PATH=","").strip()
                        if p and os.path.isdir(p):
                            return p
        except: pass
    return _SCRIPT_DIR

def save_data_path(new_path):
    lines = []
    replaced = False
    if os.path.exists(SETTINGS_FILE):
        with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
            lines = f.readlines()
        for i, line in enumerate(lines):
            if line.startswith("DATA_PATH="):
                lines[i] = f"DATA_PATH={new_path}\n"
                replaced = True
    if not replaced:
        lines.append(f"DATA_PATH={new_path}\n")
    with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
        f.writelines(lines)

def get_file(name):
    return os.path.join(load_data_path(), name)

_VISITORS_FNAME  = "GST_Visitors.xlsx"
_STAFFLOG_FNAME  = "GST_Staff_Log.xlsx"
_USERS_FNAME     = "GST_Users.xlsx"
_PHOTOS_DIRNAME  = "GST_Visitor_Photos"
_PURPOSE_FNAME   = "GST_Purposes.txt"
_OFFICERS_FNAME  = "GST_Officers.xlsx"

def VISITORS_FILE():  return get_file(_VISITORS_FNAME)
def STAFF_LOG_FILE(): return get_file(_STAFFLOG_FNAME)
def USERS_FILE():     return get_file(_USERS_FNAME)
def PURPOSE_FILE():   return get_file(_PURPOSE_FNAME)
def OFFICERS_FILE():  return get_file(_OFFICERS_FNAME)
def PHOTO_FOLDER():
    p = os.path.join(load_data_path(), _PHOTOS_DIRNAME)
    os.makedirs(p, exist_ok=True)
    return p

# Officer helpers
_OFFICER_COLS = ["Officer Code","Officer Name","Designation","Division","Room No","Block","Floor"]

def load_officers():
    fp = OFFICERS_FILE()
    if not os.path.exists(fp):
        return []
    try:
        df = pd.read_excel(fp, dtype=str).fillna("")
        for c in _OFFICER_COLS:
            if c not in df.columns:
                df[c] = ""
        return df[_OFFICER_COLS].to_dict("records")
    except:
        return []

def save_officers(records):
    pd.DataFrame(records, columns=_OFFICER_COLS).to_excel(OFFICERS_FILE(), index=False)

def find_officer(query):
    q = query.strip().lower()
    if not q:
        return None
    officers = load_officers()
    for o in officers:
        if (o.get("Officer Code","").strip().lower() == q or
                o.get("Officer Name","").strip().lower() == q):
            return o
    for o in officers:
        if q in o.get("Officer Name","").strip().lower():
            return o
    return None

def search_officers(query):
    """Return list of officers matching query (partial name or code)."""
    q = query.strip().lower()
    if not q:
        return []
    return [o for o in load_officers()
            if q in o.get("Officer Code","").lower() or
               q in o.get("Officer Name","").lower()]

_DEFAULT_PURPOSES = [
    "Official Meeting",
    "Document Submission",
    "GST Registration",
    "GST Return Filing",
    "Audit / Inspection",
    "Grievance / Complaint",
    "Certificate Collection",
    "Refund Application",
    "General Enquiry",
    "Other",
]

def load_purposes():
    pf = PURPOSE_FILE()
    if os.path.exists(pf):
        try:
            with open(pf, "r", encoding="utf-8") as f:
                items = [l.strip() for l in f if l.strip()]
            if items:
                return items
        except: pass
    return list(_DEFAULT_PURPOSES)

def save_purposes(items):
    try:
        with open(PURPOSE_FILE(), "w", encoding="utf-8") as f:
            for item in items:
                if item.strip():
                    f.write(item.strip() + "\n")
    except Exception as ex:
        print(f"Could not save purposes: {ex}")

# ═══════════════════════════════════════════
#  MODERN COLOR PALETTE — warm, professional, light
# ═══════════════════════════════════════════
# Backgrounds
BG_PAGE     = "#F5F6FA"      # warm off-white page background
BG_CARD     = "#FFFFFF"      # card/panel white
BG_HEADER   = "#1E3A5F"      # deep navy header (not dark navy, more blue)
BG_SIDEBAR  = "#F0F4FF"      # very light blue sidebar

# Accents
ACCENT_BLUE   = "#2563EB"    # primary action / links
ACCENT_GREEN  = "#16A34A"    # visitor IN / success  (muted, professional)
ACCENT_AMBER  = "#D97706"    # visitor OUT / warning  (warm amber not harsh orange)
ACCENT_RED    = "#DC2626"    # not used for buttons; reserved for errors only

# Text
TEXT_DARK   = "#1E293B"      # near-black text
TEXT_MID    = "#4B5563"      # secondary text
TEXT_LIGHT  = "#94A3B8"      # placeholder / hint text
TEXT_WHITE  = "#FFFFFF"

# Borders / dividers
BORDER_CLR  = "#E2E8F0"      # subtle border
FIELD_BG    = "#F8FAFF"      # input field background

# Status bar
STATUS_BG   = "#EFF6FF"      # light blue status bar

# Button colours (updated palette, no harsh darks)
BTN_IN      = "#16A34A"      # green — visitor in
BTN_OUT     = "#D97706"      # amber — visitor out
BTN_SEARCH  = "#2563EB"      # blue — search
BTN_REPORT  = "#7C3AED"      # soft purple — daily report
BTN_CLEAR   = "#64748B"      # slate — clear
BTN_LOGOUT_BG = "#EFF6FF"    # very light, near-white
BTN_LOGOUT_FG = "#1E3A5F"    # navy text
BTN_LOGOUT_BORDER = "#1E3A5F"

# Legacy aliases still used in a few places
NAVY   = BG_HEADER
GOLD   = "#C8A84B"
WHITE  = BG_CARD
LIGHT  = BG_PAGE
GREEN  = ACCENT_GREEN
ORANGE = ACCENT_AMBER
GRAY   = TEXT_MID
DARK   = TEXT_DARK
RED    = ACCENT_RED
PURPLE = "#7C3AED"

# ═══════════════════════════════════════════
#  INIT FOLDERS & FILES
# ═══════════════════════════════════════════
def init_files():
    vf = VISITORS_FILE()
    sl = STAFF_LOG_FILE()
    uf = USERS_FILE()
    PHOTO_FOLDER()

    if not os.path.exists(vf):
        pd.DataFrame(columns=[
            "Group ID","Date","Arrival","Out",
            "Visitor","Phone","ID Cards",
            "Total Members","Remaining",
            "Company","Purpose","GST No",
            "Officer","Division",
            "Block","Floor","Room No","Remarks","Photo"
        ]).to_excel(vf, index=False)
    else:
        df = pd.read_excel(vf, dtype=str)
        changed = False
        for col in ["Photo","Block","Floor","Room No","GST No","Officer","VIP_Code"]:
            if col not in df.columns:
                df[col] = ""; changed = True
        if "Designation" in df.columns and "Officer" not in df.columns:
            df.rename(columns={"Designation": "Officer"}, inplace=True); changed = True
        if changed:
            df.to_excel(vf, index=False)

    if not os.path.exists(sl):
        pd.DataFrame(columns=[
            "Date","Username","Full Name","Role",
            "Login Time","Logout Time","Duration (mins)"
        ]).to_excel(sl, index=False)

    if not os.path.exists(uf):
        pd.DataFrame({
            "Username":  ["admin",          "reception",        "supervisor"],
            "Password":  ["admin123",       "123",              "super456"],
            "Full Name": ["Administrator",  "Reception Officer","Supervisor"],
            "Role":      ["admin",          "staff",            "staff"],
            "Mobile":    ["9999999999",     "8888888888",       "7777777777"]
        }).to_excel(uf, index=False)
    else:
        # Auto-migrate: add Mobile column if missing
        try:
            udf = pd.read_excel(uf, dtype=str)
            if "Mobile" not in udf.columns:
                udf["Mobile"] = ""
                udf.to_excel(uf, index=False)
        except: pass

def first_run_setup():
    if os.path.exists(SETTINGS_FILE):
        return
    if not getattr(_sys, "frozen", False):
        return
    import tkinter as _tk
    from tkinter import messagebox as _mb, filedialog as _fd
    _r = _tk.Tk(); _r.withdraw()
    _mb.showinfo(
        "GST Visitor System — First Time Setup",
        "Welcome!\n\nPlease choose a folder where all your visitor data\n"
        "will be saved permanently.\n\nRecommended:\n"
        "  C:\\Users\\YourName\\Documents\\GST_VMS_Data\n\n"
        "Avoid Desktop or Downloads (can be accidentally deleted).\n"
        "Avoid any Temp or AppData folder.",
        parent=_r)
    chosen = _fd.askdirectory(
        title="Choose Data Save Folder for GST Visitor System",
        initialdir=os.path.expanduser("~\\Documents"))
    if chosen:
        chosen = chosen.replace("/", "\\")
        os.makedirs(chosen, exist_ok=True)
        save_data_path(chosen)
        _mb.showinfo("✅  Folder Set",
            f"Data will be saved to:\n{chosen}\n\n"
            "You can change this later in\nAdmin Login → Settings tab.", parent=_r)
    else:
        docs = os.path.join(os.path.expanduser("~"), "Documents", "GST_VMS_Data")
        os.makedirs(docs, exist_ok=True)
        save_data_path(docs)
        _mb.showinfo("✅  Default Folder Set",
            f"No folder chosen. Using default:\n{docs}\n\n"
            "You can change this in Admin Login → Settings tab.", parent=_r)
    _r.destroy()

first_run_setup()
init_files()

# ═══════════════════════════════════════════
#  AUTO BACKUP
# ═══════════════════════════════════════════
def auto_backup():
    folder = os.path.join(load_data_path(), "GST_Backups")
    os.makedirs(folder, exist_ok=True)
    today = datetime.now().strftime(DATE_FORMAT)
    for f in [VISITORS_FILE(), STAFF_LOG_FILE()]:
        if os.path.exists(f):
            bname = os.path.basename(f).replace('.xlsx','')
            bk = os.path.join(folder, f"{bname}_Backup_{today}.xlsx")
            if not os.path.exists(bk):
                shutil.copy2(f, bk)

auto_backup()

# ═══════════════════════════════════════════
#  SESSION
# ═══════════════════════════════════════════
session = {"username": None, "full_name": None,
           "role": None, "login_time": None}

def verify_login(user, pwd):
    try:
        df = pd.read_excel(USERS_FILE(), dtype=str)
        df["Username"] = df["Username"].str.strip()
        df["Password"] = df["Password"].str.strip()
        m = df[(df["Username"]==user.strip()) & (df["Password"]==pwd.strip())]
        if not m.empty:
            r = m.iloc[0]
            return {"username": r["Username"],
                    "full_name": r["Full Name"],
                    "role": r["Role"]}
    except Exception as e:
        messagebox.showerror("Error", str(e))
    return None

def log_login(u, fn, role):
    session["username"]   = u
    session["full_name"]  = fn
    session["role"]       = role
    session["login_time"] = datetime.now()

def log_logout():
    if not session["login_time"]: return
    now = datetime.now()
    dur = round((now - session["login_time"]).total_seconds()/60, 2)
    row = {
        "Date":            session["login_time"].strftime(DATE_FORMAT),
        "Username":        session["username"],
        "Full Name":       session["full_name"],
        "Role":            session["role"],
        "Login Time":      session["login_time"].strftime(TIME_FORMAT),
        "Logout Time":     now.strftime(TIME_FORMAT),
        "Duration (mins)": dur
    }
    df = pd.read_excel(STAFF_LOG_FILE(), dtype=str)
    df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
    df.to_excel(STAFF_LOG_FILE(), index=False)
    for k in session: session[k] = None

# ═══════════════════════════════════════════
#  USB BACKUP
# ═══════════════════════════════════════════
def backup_to_usb(parent=None):
    found = False
    for drv in ["D:","E:","F:","G:","H:"]:
        if os.path.exists(drv + "\\"):
            today = datetime.now().strftime(DATE_FORMAT).replace("-","_")
            backed = []
            for f in [VISITORS_FILE(), STAFF_LOG_FILE()]:
                if os.path.exists(f):
                    bname = os.path.basename(f).replace('.xlsx','')
                    dst = f"{drv}\\GST_{bname}_{today}.xlsx"
                    shutil.copy2(f, dst)
                    backed.append(dst)
            messagebox.showinfo("USB Backup Done",
                "Backed up to " + drv + "\n\n" + "\n".join(backed))
            found = True; break
    if not found:
        messagebox.showwarning("No USB",
            "No USB drive found.\nData is also backed up in GST_Backups\\ folder.")

# ═══════════════════════════════════════════
#  DESKTOP SHORTCUT
# ═══════════════════════════════════════════
def create_desktop_shortcut():
    try:
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        sc = os.path.join(desktop, "GST Visitor System.lnk")
        if os.path.exists(sc): return
        import sys
        pyexe = sys.executable.replace("python.exe","pythonw.exe")
        if not os.path.exists(pyexe): pyexe = sys.executable
        script = os.path.abspath(__file__)
        ps = (f'$s=New-Object -comObject WScript.Shell\n'
              f'$l=$s.CreateShortcut("{sc}")\n'
              f'$l.TargetPath="{pyexe}"\n'
              f'$l.Arguments=\'"{script}"\'\n'
              f'$l.WorkingDirectory="{os.path.dirname(script)}"\n'
              f'$l.Description="GST Visitor System"\n'
              f'$l.Save()')
        psf = os.path.join(os.path.dirname(script), "_sc.ps1")
        with open(psf,"w") as f: f.write(ps)
        import subprocess
        subprocess.run(["powershell","-ExecutionPolicy","Bypass","-File",psf],
                       capture_output=True, timeout=10)
        if os.path.exists(psf): os.remove(psf)
        if os.path.exists(sc):
            messagebox.showinfo("Shortcut Created",
                "GST Visitor System shortcut added to Desktop!")
    except: pass

create_desktop_shortcut()

# ═══════════════════════════════════════════
#  WEBCAM — DETECT & CAPTURE
# ═══════════════════════════════════════════
def detect_webcam():
    if not WEBCAM_AVAILABLE:
        return -1
    for index in range(3):
        cap = cv2.VideoCapture(index, cv2.CAP_DSHOW)
        if cap.isOpened():
            ret, _ = cap.read()
            cap.release()
            if ret:
                return index
    return -1

def capture_visitor_photo(group_id, visitor_name, parent=None):
    if not WEBCAM_AVAILABLE:
        messagebox.showwarning("Camera Library Missing",
            "Webcam feature needs extra libraries.\n\n"
            "Run this once in Command Prompt:\n"
            "pip install opencv-python pillow\n\n"
            "Then restart the system.", parent=parent)
        return ""

    cam_index = detect_webcam()
    result = {"path": ""}

    popup = Toplevel(parent)
    popup.title(f"Capture Photo — {visitor_name}")
    popup.geometry("520x480")
    popup.configure(bg=BG_PAGE)
    popup.grab_set()
    popup.resizable(False, False)

    # Header
    hdr = Frame(popup, bg=BG_HEADER, pady=12)
    hdr.pack(fill=X)
    Label(hdr, text="📷  Visitor Photo Capture",
          font=("Segoe UI", 11, "bold"), bg=BG_HEADER, fg=TEXT_WHITE).pack()
    Label(hdr, text=visitor_name,
          font=("Segoe UI", 9), bg=BG_HEADER, fg="#93C5FD").pack()

    card = Frame(popup, bg=BG_CARD, bd=0, relief=FLAT,
                 highlightthickness=1, highlightbackground=BORDER_CLR)
    card.pack(fill=BOTH, expand=True, padx=16, pady=12)

    if cam_index < 0:
        Label(card, text="No webcam detected",
              font=("Segoe UI", 11), bg=BG_CARD, fg=TEXT_MID).pack(expand=True)
        Button(card, text="Close", font=("Segoe UI", 10, "bold"),
               bg=BTN_CLEAR, fg=TEXT_WHITE, relief=FLAT, cursor="hand2",
               padx=20, pady=8, command=popup.destroy).pack(pady=12)
        popup.wait_window()
        return ""

    cap = cv2.VideoCapture(cam_index, cv2.CAP_DSHOW)
    cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)

    vid_lbl = Label(card, bg=BG_CARD)
    vid_lbl.pack(padx=8, pady=(8,4))

    info_lbl = Label(card, text="Position visitor in frame, then click Capture",
                     font=("Segoe UI", 9), bg=BG_CARD, fg=TEXT_MID)
    info_lbl.pack()

    btn_row = Frame(card, bg=BG_CARD); btn_row.pack(pady=8)
    running = [True]

    def stream():
        if not running[0]: return
        ret, frame = cap.read()
        if ret:
            frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            img = Image.fromarray(frame_rgb).resize((380, 285), Image.LANCZOS)
            imgtk = ImageTk.PhotoImage(image=img)
            vid_lbl.imgtk = imgtk
            vid_lbl.config(image=imgtk)
        popup.after(33, stream)

    stream()

    def do_capture():
        ret, frame = cap.read()
        if not ret:
            messagebox.showerror("Error", "Could not capture frame.", parent=popup); return
        grp_folder = os.path.join(PHOTO_FOLDER(), group_id)
        os.makedirs(grp_folder, exist_ok=True)
        ts   = datetime.now().strftime("%H%M%S")
        cnt  = len([f for f in os.listdir(grp_folder) if f.endswith(".jpg")]) + 1
        fn   = os.path.join(grp_folder, f"{cnt}_{ts}.jpg")
        cv2.imwrite(fn, frame)
        result["path"] = fn
        info_lbl.config(text=f"✅  Photo saved!", fg=ACCENT_GREEN)
        messagebox.showinfo("Photo Saved",
            f"Visitor  : {visitor_name}\nSaved   : {fn}", parent=popup)
        running[0] = False
        cap.release()
        popup.destroy()

    def skip():
        running[0] = False
        cap.release()
        popup.destroy()

    Button(btn_row, text="📷  Capture",
           font=("Segoe UI", 10, "bold"), bg=BTN_IN, fg=TEXT_WHITE,
           relief=FLAT, cursor="hand2", padx=18, pady=8,
           command=do_capture).pack(side=LEFT, padx=6)
    Button(btn_row, text="Skip",
           font=("Segoe UI", 10), bg=BG_PAGE, fg=TEXT_MID,
           relief=FLAT, cursor="hand2", padx=18, pady=8,
           highlightthickness=1, highlightbackground=BORDER_CLR,
           command=skip).pack(side=LEFT, padx=6)

    def on_close():
        running[0] = False
        try: cap.release()
        except: pass
        popup.destroy()
    popup.protocol("WM_DELETE_WINDOW", on_close)
    popup.wait_window()
    return result["path"]

# ═══════════════════════════════════════════
#  VIEW SAVED PHOTO
# ═══════════════════════════════════════════
def view_visitor_photo(photo_path, visitor_name, parent=None):
    if not WEBCAM_AVAILABLE:
        messagebox.showinfo("Library Missing",
            "Install pillow to view photos:\npip install pillow",
            parent=parent); return

    actual_file = ""
    grp_folder  = ""

    if photo_path and os.path.isfile(photo_path):
        actual_file = photo_path
        grp_folder  = os.path.dirname(photo_path)
    elif photo_path and os.path.isdir(photo_path):
        grp_folder = photo_path
        jpgs = [f for f in os.listdir(grp_folder) if f.lower().endswith(".jpg")]
        if jpgs:
            actual_file = os.path.join(grp_folder, jpgs[0])
    else:
        messagebox.showinfo("No Photo",
            f"No photo on record for {visitor_name}.\n\nPath stored: {photo_path}",
            parent=parent)
        return

    popup = Toplevel(parent)
    popup.title(f"Photo — {visitor_name}")
    popup.geometry("380x460")
    popup.configure(bg=BG_PAGE)
    popup.grab_set()
    popup.resizable(False, False)

    hdr = Frame(popup, bg=BG_HEADER, pady=12)
    hdr.pack(fill=X)
    Label(hdr, text=f"📷  {visitor_name}",
          font=("Segoe UI", 11, "bold"), bg=BG_HEADER, fg=TEXT_WHITE).pack()
    grp_name = os.path.basename(grp_folder) if grp_folder else ""
    Label(hdr, text=f"Group: {grp_name}",
          font=("Segoe UI", 8), bg=BG_HEADER, fg="#93C5FD").pack()

    card = Frame(popup, bg=BG_CARD, bd=0, relief=FLAT,
                 highlightthickness=1, highlightbackground=BORDER_CLR)
    card.pack(fill=BOTH, expand=True, padx=16, pady=12)

    if actual_file and os.path.isfile(actual_file):
        img   = Image.open(actual_file).resize((300, 225), Image.LANCZOS)
        imgtk = ImageTk.PhotoImage(image=img)
        lbl   = Label(card, image=imgtk, bg=BG_CARD)
        lbl.imgtk = imgtk; lbl.pack(pady=(12,4))
        sz = os.path.getsize(actual_file)
        Label(card, text=f"{os.path.basename(actual_file)}  •  {sz} bytes",
              font=("Segoe UI", 8), bg=BG_CARD, fg=TEXT_LIGHT).pack()
    else:
        Label(card, text="(image file not found)",
              font=("Segoe UI", 9, "italic"), bg=BG_CARD, fg=TEXT_LIGHT).pack(pady=30)

    btn_f = Frame(card, bg=BG_CARD, pady=10); btn_f.pack()

    def open_folder():
        if grp_folder and os.path.isdir(grp_folder):
            os.startfile(grp_folder)
        else:
            messagebox.showinfo("Not Found", "Photo folder not found.", parent=popup)

    Button(btn_f, text="📂  Open Folder",
           font=("Segoe UI", 9, "bold"), bg=BTN_OUT, fg=TEXT_WHITE,
           relief=FLAT, cursor="hand2", padx=12, pady=7,
           command=open_folder).pack(side=LEFT, padx=6)
    Button(btn_f, text="Close",
           font=("Segoe UI", 9), bg=BG_PAGE, fg=TEXT_MID,
           relief=FLAT, cursor="hand2", padx=12, pady=7,
           highlightthickness=1, highlightbackground=BORDER_CLR,
           command=popup.destroy).pack(side=LEFT, padx=6)

# ═══════════════════════════════════════════
#  SHARED UI HELPERS
# ═══════════════════════════════════════════
def apply_modern_style():
    """Apply ttk styles globally for a clean, modern look."""
    style = ttk.Style()
    style.theme_use("clam")

    # Treeview
    style.configure("Modern.Treeview",
        background=BG_CARD, foreground=TEXT_DARK,
        fieldbackground=BG_CARD, font=("Segoe UI", 9),
        rowheight=26, borderwidth=0)
    style.configure("Modern.Treeview.Heading",
        font=("Segoe UI", 9, "bold"),
        background=BG_HEADER, foreground=TEXT_WHITE,
        borderwidth=0, relief="flat")
    style.map("Modern.Treeview",
        background=[("selected", "#DBEAFE")],
        foreground=[("selected", TEXT_DARK)])
    style.map("Modern.Treeview.Heading",
        background=[("active", BG_HEADER)])

    # Combobox
    style.configure("Modern.TCombobox",
        fieldbackground=FIELD_BG, background=FIELD_BG,
        foreground=TEXT_DARK, font=("Segoe UI", 10),
        bordercolor=BORDER_CLR, arrowcolor=TEXT_MID,
        padding=(8, 6))
    style.map("Modern.TCombobox",
        fieldbackground=[("readonly", FIELD_BG)],
        bordercolor=[("focus", ACCENT_BLUE)])

    # Scrollbar — thicker and visible
    style.configure("Modern.Vertical.TScrollbar",
        background="#94A3B8", troughcolor="#E2E8F0", width=14,
        borderwidth=0, arrowsize=12, relief="flat")
    style.configure("Modern.Horizontal.TScrollbar",
        background="#94A3B8", troughcolor="#E2E8F0", width=14,
        borderwidth=0, arrowsize=12, relief="flat")
    style.map("Modern.Vertical.TScrollbar",
        background=[("active","#64748B"),("pressed","#475569")])

def make_card(parent, **kwargs):
    """Returns a white rounded-style card frame."""
    defaults = dict(bg=BG_CARD, bd=0, relief=FLAT,
                    highlightthickness=1, highlightbackground=BORDER_CLR)
    defaults.update(kwargs)
    return Frame(parent, **defaults)

def make_entry(parent, width=28, textvariable=None, show=None):
    """Returns a styled flat entry widget."""
    kw = dict(
        font=("Segoe UI", 10), bg=FIELD_BG, fg=TEXT_DARK,
        relief=FLAT, highlightthickness=1,
        highlightbackground=BORDER_CLR, highlightcolor=ACCENT_BLUE,
        insertbackground=ACCENT_BLUE, width=width, bd=0)
    if textvariable: kw["textvariable"] = textvariable
    if show:         kw["show"] = show
    e = Entry(parent, **kw)
    # Hover effect
    e.bind("<FocusIn>",  lambda ev, w=e: w.config(highlightbackground=ACCENT_BLUE,
                                                    highlightcolor=ACCENT_BLUE, bg="#F0F7FF"))
    e.bind("<FocusOut>", lambda ev, w=e: w.config(highlightbackground=BORDER_CLR,
                                                   highlightcolor=ACCENT_BLUE, bg=FIELD_BG))
    return e

def make_button(parent, text, bg, fg=TEXT_WHITE, command=None,
                padx=18, pady=8, width=None, font_size=9):
    """Returns a clean, flat button with hover effect."""
    import colorsys
    def darken(hex_col, factor=0.85):
        try:
            h = hex_col.lstrip("#")
            r, g, b = int(h[0:2],16)/255, int(h[2:4],16)/255, int(h[4:6],16)/255
            r2, g2, b2 = r*factor, g*factor, b*factor
            return "#{:02x}{:02x}{:02x}".format(int(r2*255),int(g2*255),int(b2*255))
        except:
            return hex_col
    kw = dict(
        text=text, bg=bg, fg=fg,
        font=("Segoe UI", font_size, "bold"),
        relief=FLAT, cursor="hand2",
        activebackground=darken(bg),
        activeforeground=fg,
        padx=padx, pady=pady, bd=0)
    if command: kw["command"] = command
    if width:   kw["width"] = width
    btn = Button(parent, **kw)
    return btn

def build_header(root, title, subtitle=""):
    """Modern header with logo, title, clock, user badge."""
    hdr = Frame(root, bg=BG_HEADER)
    hdr.pack(fill=X)

    inner = Frame(hdr, bg=BG_HEADER, pady=14)
    inner.pack(fill=X, padx=20)

    # Left — logo + title block
    lf = Frame(inner, bg=BG_HEADER); lf.pack(side=LEFT)

    logo_box = Frame(lf, bg="#2D5282", width=50, height=50,
                     highlightthickness=1, highlightbackground="#4A7CBC")
    logo_box.pack(side=LEFT, padx=(0,14))
    logo_box.pack_propagate(False)
    Label(logo_box, text="🏛", font=("Segoe UI Emoji", 26),
          bg="#2D5282", fg="#F6D860").place(relx=0.5, rely=0.5, anchor=CENTER)

    tf = Frame(lf, bg=BG_HEADER); tf.pack(side=LEFT)
    Label(tf, text="ಕರ್ನಾಟಕ ಸರ್ಕಾರ  •  GOVERNMENT OF KARNATAKA",
          font=("Nirmala UI", 8, "bold"), bg=BG_HEADER, fg="#93C5FD").pack(anchor=W)
    Label(tf, text=title,
          font=("Segoe UI", 15, "bold"), bg=BG_HEADER, fg=TEXT_WHITE).pack(anchor=W)
    Label(tf, text="Goods and Services Tax Department  •  ಸತ್ಯಮೇವ ಜಯತೇ",
          font=("Segoe UI", 8), bg=BG_HEADER, fg="#64A8D4").pack(anchor=W)

    # Right — clock + user badge
    rf = Frame(inner, bg=BG_HEADER); rf.pack(side=RIGHT)
    tv = StringVar(); dv = StringVar()
    def clk():
        n = datetime.now()
        tv.set(n.strftime("%H:%M:%S"))
        dv.set(n.strftime("%d %B %Y"))
        root.after(1000, clk)
    Label(rf, textvariable=dv, font=("Segoe UI", 9),
          bg=BG_HEADER, fg="#93C5FD").pack(anchor=E)
    Label(rf, textvariable=tv, font=("Segoe UI", 14, "bold"),
          bg=BG_HEADER, fg=TEXT_WHITE).pack(anchor=E)

    # User badge pill
    uname = session.get("full_name","User").upper()
    badge = Frame(rf, bg="#2D5282", bd=0,
                  highlightthickness=1, highlightbackground="#4A7CBC")
    badge.pack(anchor=E, pady=(4,0))
    Label(badge, text=f"  👤  {uname}  ",
          font=("Segoe UI", 8, "bold"), bg="#2D5282", fg="#E0F2FE",
          pady=4).pack()

    clk()

    # Thin accent line at bottom of header
    Frame(hdr, bg="#3B82F6", height=3).pack(fill=X)
    return hdr

def build_treeview(parent, columns, height=8):
    fr = Frame(parent, bg=BG_PAGE)
    fr.pack(fill=BOTH, expand=True, padx=10, pady=6)
    tree = ttk.Treeview(fr, columns=columns,
                        show="headings", style="Modern.Treeview", height=height)
    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, width=95, anchor=CENTER)
    sby = ttk.Scrollbar(fr, orient=VERTICAL,   command=tree.yview,
                        style="Modern.Vertical.TScrollbar")
    sbx = ttk.Scrollbar(fr, orient=HORIZONTAL, command=tree.xview,
                        style="Modern.Horizontal.TScrollbar")
    tree.configure(yscrollcommand=sby.set, xscrollcommand=sbx.set)
    sby.pack(side=RIGHT, fill=Y)
    sbx.pack(side=BOTTOM, fill=X)
    tree.pack(fill=BOTH, expand=True)
    # Alternating row colours
    tree.tag_configure("odd",  background="#F8FAFF")
    tree.tag_configure("even", background=BG_CARD)
    return tree

def insert_tree_rows(tree, df):
    for i, (_, r) in enumerate(df.iterrows()):
        tag = "odd" if i % 2 == 0 else "even"
        tree.insert("", END, values=list(r), tags=(tag,))

def build_bottom_bar(root, logout_cmd):
    Frame(root, bg=BORDER_CLR, height=1).pack(fill=X, side=BOTTOM)
    bar = Frame(root, bg=BG_PAGE, pady=10)
    bar.pack(fill=X, side=BOTTOM)
    make_button(bar, "  💾  USB Backup", BTN_OUT,
                command=lambda: backup_to_usb(root),
                padx=14, pady=8).pack(side=RIGHT, padx=10)
    # Logout — outlined style
    lo_btn = Button(bar, text="  🔒  Logout  ",
                    font=("Segoe UI", 9, "bold"),
                    bg=BTN_LOGOUT_BG, fg=BTN_LOGOUT_FG,
                    relief=FLAT, cursor="hand2",
                    activebackground="#DBEAFE", activeforeground=BG_HEADER,
                    padx=14, pady=8, bd=0,
                    highlightthickness=1, highlightbackground=BTN_LOGOUT_BORDER,
                    command=logout_cmd)
    lo_btn.pack(side=RIGHT, padx=4)

# ═══════════════════════════════════════════
#  LOGOUT
# ═══════════════════════════════════════════
def do_logout(win):
    name = session.get("full_name","User")
    lt   = session.get("login_time")
    if lt:
        dur = round((datetime.now()-lt).total_seconds()/60, 2)
        msg = ("Logging out: " + name + "\n\n"
               "Login  : " + lt.strftime(TIME_FORMAT) + "\n"
               "Logout : " + datetime.now().strftime(TIME_FORMAT) + "\n"
               "Duration: " + str(dur) + " minutes")
        log_logout(); win.destroy()
        messagebox.showinfo("Logged Out", msg)
    else:
        log_logout(); win.destroy()
    show_login()

# ═══════════════════════════════════════════
#  LOGIN WINDOW — matches image: blue gradient bg, clean white card
# ═══════════════════════════════════════════
def show_login():
    lw = Tk()
    lw.title("GST Visitor Management System")
    lw.geometry("480x560")
    lw.resizable(False, False)
    apply_modern_style()

    # ── Blue gradient background (matches image 2) ──
    # Simulated with a canvas gradient
    bg_canvas = Canvas(lw, highlightthickness=0)
    bg_canvas.pack(fill=BOTH, expand=True)

    def _draw_bg(e=None):
        w = bg_canvas.winfo_width() or 480
        h = bg_canvas.winfo_height() or 560
        bg_canvas.delete("all")
        # Draw soft blue gradient: top #D6E8FF → bottom #EEF4FF
        steps = 40
        for i in range(steps):
            ratio = i / steps
            r = int(0xD6 + (0xEE - 0xD6) * ratio)
            g = int(0xE8 + (0xF4 - 0xE8) * ratio)
            b = int(0xFF)
            color = f"#{r:02x}{g:02x}{b:02x}"
            bg_canvas.create_rectangle(0, i*(h//steps), w,
                                       (i+1)*(h//steps)+2,
                                       fill=color, outline="")
        # Soft blobs for depth (like image 2)
        bg_canvas.create_oval(-60, -60, 200, 200,
                              fill="#C7DCFF", outline="")
        bg_canvas.create_oval(300, 320, 560, 580,
                              fill="#D4E9FF", outline="")
        _build_card()

    # ── Build card on top of canvas ──
    _card_built = [False]
    def _build_card():
        if _card_built[0]: return
        _card_built[0] = True

        # White card centered
        card = Frame(bg_canvas, bg="#FFFFFF", bd=0,
                     highlightthickness=1, highlightbackground="#D1D9E6")
        card.place(relx=0.5, rely=0.5, anchor=CENTER,
                   width=360, height=430)

        inner = Frame(card, bg="#FFFFFF", padx=32, pady=28)
        inner.pack(fill=BOTH, expand=True)

        # Icon top
        ic_f = Frame(inner, bg="#FFFFFF"); ic_f.pack()
        ic_cv = Canvas(ic_f, width=56, height=56, bg="#FFFFFF",
                       highlightthickness=0)
        ic_cv.pack()
        ic_cv.create_oval(4, 4, 52, 52, fill="#EFF6FF",
                          outline="#BFDBFE", width=1.5)
        ic_cv.create_text(28, 28, text="🏛", font=("Segoe UI Emoji", 24),
                          fill="#1E3A5F")

        # Title
        Label(inner, text="Welcome Back",
              font=("Segoe UI", 18, "bold"),
              bg="#FFFFFF", fg="#1E293B").pack(pady=(12, 2))
        Label(inner, text="Please sign in to continue.",
              font=("Segoe UI", 9),
              bg="#FFFFFF", fg="#94A3B8").pack(pady=(0, 20))

        # Username field
        uv = StringVar()
        uf_frame = Frame(inner, bg="#FFFFFF",
                         highlightthickness=1,
                         highlightbackground="#E2E8F0")
        uf_frame.pack(fill=X, pady=(0, 10))
        uf_in = Frame(uf_frame, bg="#F8FAFF", padx=10, pady=0)
        uf_in.pack(fill=X)
        Label(uf_in, text="👤", font=("Segoe UI Emoji", 11),
              bg="#F8FAFF", fg="#94A3B8").pack(side=LEFT, pady=10)
        ue = Entry(uf_in, textvariable=uv, font=("Segoe UI", 10),
                   bg="#F8FAFF", fg="#1E293B", relief=FLAT, bd=0,
                   highlightthickness=0, insertbackground="#2563EB")
        ue.insert(0, "Username")
        ue.config(fg="#94A3B8")
        ue.pack(side=LEFT, fill=X, expand=True, pady=10, padx=6)

        def _ue_in(e):
            if ue.get() == "Username":
                ue.delete(0, END); ue.config(fg="#1E293B")
            uf_frame.config(highlightbackground="#2563EB")
        def _ue_out(e):
            if not ue.get(): ue.insert(0,"Username"); ue.config(fg="#94A3B8")
            uf_frame.config(highlightbackground="#E2E8F0")
        ue.bind("<FocusIn>",  _ue_in)
        ue.bind("<FocusOut>", _ue_out)

        # Password field
        pv = StringVar()
        pf_frame = Frame(inner, bg="#FFFFFF",
                         highlightthickness=1,
                         highlightbackground="#E2E8F0")
        pf_frame.pack(fill=X, pady=(0, 6))
        pf_in = Frame(pf_frame, bg="#F8FAFF", padx=10, pady=0)
        pf_in.pack(fill=X)
        Label(pf_in, text="🔒", font=("Segoe UI Emoji", 11),
              bg="#F8FAFF", fg="#94A3B8").pack(side=LEFT, pady=10)
        show_var = BooleanVar(value=False)
        pe = Entry(pf_in, textvariable=pv, font=("Segoe UI", 10),
                   bg="#F8FAFF", fg="#1E293B", relief=FLAT, bd=0,
                   highlightthickness=0, insertbackground="#2563EB",
                   show="")
        pe.insert(0, "Password")
        pe.config(fg="#94A3B8")
        pe.pack(side=LEFT, fill=X, expand=True, pady=10, padx=6)

        _pw_has_placeholder = [True]
        def _pe_in(e):
            if _pw_has_placeholder[0]:
                pe.delete(0,END); pe.config(fg="#1E293B", show="●")
                _pw_has_placeholder[0] = False
            pf_frame.config(highlightbackground="#2563EB")
        def _pe_out(e):
            if not pe.get():
                pe.config(show=""); pe.insert(0,"Password")
                pe.config(fg="#94A3B8"); _pw_has_placeholder[0] = True
            pf_frame.config(highlightbackground="#E2E8F0")
        pe.bind("<FocusIn>",  _pe_in)
        pe.bind("<FocusOut>", _pe_out)

        # Remember me row
        row_f = Frame(inner, bg="#FFFFFF"); row_f.pack(fill=X, pady=(4, 16))
        rm_var = BooleanVar()
        Checkbutton(row_f, text="Remember me", variable=rm_var,
                    bg="#FFFFFF", fg="#4B5563",
                    activebackground="#FFFFFF",
                    selectcolor="#FFFFFF",
                    font=("Segoe UI", 8), cursor="hand2").pack(side=LEFT)

        # Forgot password link
        fp_btn = Label(row_f, text="Forgot password?",
                       font=("Segoe UI", 8, "underline"),
                       bg="#FFFFFF", fg="#2563EB", cursor="hand2")
        fp_btn.pack(side=RIGHT)
        fp_btn.bind("<Button-1>", lambda e: _show_forgot())

        # Login button — full width, primary blue
        def do_login():
            u = uv.get().strip()
            p = pv.get().strip() if not _pw_has_placeholder[0] else ""
            if u == "Username": u = ""
            if not u or not p:
                messagebox.showwarning("Missing",
                    "Enter username and password.", parent=lw); return
            res = verify_login(u, p)
            if res:
                log_login(res["username"], res["full_name"], res["role"])
                lw.destroy()
                _dp = load_data_path()
                _dl = _dp.lower()
                if ("\\temp\\" in _dl or "appdata\\local\\temp" in _dl):
                    messagebox.showwarning("⚠️  Temp Folder Warning",
                        f"Data is saving to a TEMP folder:\n{_dp}\n\n"
                        "Please change it in Settings tab.")
                if res["role"] == "admin": open_admin()
                else: open_reception()
            else:
                messagebox.showerror("Access Denied",
                    "Invalid username or password.", parent=lw)

        login_btn = Button(inner, text="Log In",
                           font=("Segoe UI", 11, "bold"),
                           bg="#2563EB", fg="#FFFFFF",
                           relief=FLAT, cursor="hand2",
                           activebackground="#1D4ED8",
                           activeforeground="#FFFFFF",
                           pady=11, bd=0)
        login_btn.config(command=do_login)
        login_btn.pack(fill=X)

        # Bottom info
        Label(card, text="Authorised Personnel Only  •  GST Department",
              font=("Segoe UI", 7), bg="#FFFFFF",
              fg="#CBD5E1").pack(pady=(0, 4))
        Label(card, text="Developed by: Udbhav K  •  v15",
              font=("Segoe UI", 7), bg="#FFFFFF",
              fg="#CBD5E1").pack(pady=(0, 12))

    # ── FORGOT PASSWORD WINDOW ──
    def _show_forgot():
        fw = Toplevel(lw)
        fw.title("Reset Password")
        fw.geometry("400x400")
        fw.resizable(False, False)
        fw.configure(bg="#F5F6FA")
        fw.grab_set()

        # Card
        fc = Frame(fw, bg="#FFFFFF",
                   highlightthickness=1, highlightbackground="#E2E8F0")
        fc.place(relx=0.5, rely=0.5, anchor=CENTER, width=340, height=360)
        fi = Frame(fc, bg="#FFFFFF", padx=28, pady=24)
        fi.pack(fill=BOTH, expand=True)

        Label(fi, text="🔐", font=("Segoe UI Emoji", 28),
              bg="#FFFFFF").pack()
        Label(fi, text="Reset Password",
              font=("Segoe UI", 15, "bold"),
              bg="#FFFFFF", fg="#1E293B").pack(pady=(8, 4))
        Label(fi, text="Enter your Username and registered\nMobile Number to verify.",
              font=("Segoe UI", 9), bg="#FFFFFF",
              fg="#94A3B8", justify=CENTER).pack(pady=(0, 16))

        # Step 1: verify username + mobile
        step1 = Frame(fi, bg="#FFFFFF"); step1.pack(fill=X)
        step2 = Frame(fi, bg="#FFFFFF")  # hidden initially

        # ── Step 1 fields ──
        Label(step1, text="Username", font=("Segoe UI",8,"bold"),
              bg="#FFFFFF", fg="#4B5563").pack(anchor=W)
        fu_var = StringVar()
        fu_e = Entry(step1, textvariable=fu_var, font=("Segoe UI",10),
                     bg="#F8FAFF", fg="#1E293B", relief=FLAT,
                     highlightthickness=1, highlightbackground="#E2E8F0",
                     highlightcolor="#2563EB", bd=0, insertbackground="#2563EB")
        fu_e.pack(fill=X, ipady=8, pady=(3,12))

        Label(step1, text="Registered Mobile Number",
              font=("Segoe UI",8,"bold"), bg="#FFFFFF", fg="#4B5563").pack(anchor=W)
        fm_var = StringVar()
        fm_e = Entry(step1, textvariable=fm_var, font=("Segoe UI",10),
                     bg="#F8FAFF", fg="#1E293B", relief=FLAT,
                     highlightthickness=1, highlightbackground="#E2E8F0",
                     highlightcolor="#2563EB", bd=0, insertbackground="#2563EB")
        fm_e.pack(fill=X, ipady=8, pady=(3,16))

        status_lbl = Label(step1, text="", font=("Segoe UI",8),
                           bg="#FFFFFF", fg="#DC2626")
        status_lbl.pack(anchor=W, pady=(0,8))

        def verify_step1():
            u = fu_var.get().strip()
            m = fm_var.get().strip()
            if not u or not m:
                status_lbl.config(text="Please fill both fields.", fg="#DC2626"); return
            try:
                df = pd.read_excel(USERS_FILE(), dtype=str)
                df["Username"] = df["Username"].str.strip()
                df["Mobile"]   = df["Mobile"].astype(str).str.strip() if "Mobile" in df.columns else pd.Series([""] * len(df))
                match = df[(df["Username"]==u) & (df["Mobile"]==m)]
                if match.empty:
                    status_lbl.config(
                        text="❌  Username or Mobile not found.", fg="#DC2626")
                    return
                status_lbl.config(
                    text="✅  Verified! Set your new password.", fg="#16A34A")
                # Show step 2
                step1.pack_forget()
                step2.pack(fill=X)
            except Exception as ex:
                status_lbl.config(text=f"Error: {ex}", fg="#DC2626")

        Button(step1, text="Verify",
               font=("Segoe UI",10,"bold"),
               bg="#2563EB", fg="#FFFFFF",
               relief=FLAT, cursor="hand2",
               activebackground="#1D4ED8",
               pady=9, bd=0,
               command=verify_step1).pack(fill=X)

        # ── Step 2: new password ──
        Label(step2, text="New Password",
              font=("Segoe UI",8,"bold"),
              bg="#FFFFFF", fg="#4B5563").pack(anchor=W)
        np_var = StringVar()
        np_e = Entry(step2, textvariable=np_var, font=("Segoe UI",10),
                     bg="#F8FAFF", fg="#1E293B", relief=FLAT, show="●",
                     highlightthickness=1, highlightbackground="#E2E8F0",
                     highlightcolor="#2563EB", bd=0, insertbackground="#2563EB")
        np_e.pack(fill=X, ipady=8, pady=(3,12))

        Label(step2, text="Confirm New Password",
              font=("Segoe UI",8,"bold"),
              bg="#FFFFFF", fg="#4B5563").pack(anchor=W)
        cp_var = StringVar()
        cp_e = Entry(step2, textvariable=cp_var, font=("Segoe UI",10),
                     bg="#F8FAFF", fg="#1E293B", relief=FLAT, show="●",
                     highlightthickness=1, highlightbackground="#E2E8F0",
                     highlightcolor="#2563EB", bd=0, insertbackground="#2563EB")
        cp_e.pack(fill=X, ipady=8, pady=(3,16))

        s2_lbl = Label(step2, text="", font=("Segoe UI",8),
                       bg="#FFFFFF", fg="#DC2626")
        s2_lbl.pack(anchor=W, pady=(0,8))

        def save_new_password():
            np_ = np_var.get().strip()
            cp_ = cp_var.get().strip()
            if not np_ or not cp_:
                s2_lbl.config(text="Please fill both fields.", fg="#DC2626"); return
            if np_ != cp_:
                s2_lbl.config(text="❌  Passwords do not match.", fg="#DC2626"); return
            if len(np_) < 4:
                s2_lbl.config(text="❌  Password must be at least 4 characters.", fg="#DC2626"); return
            try:
                df = pd.read_excel(USERS_FILE(), dtype=str)
                df["Username"] = df["Username"].str.strip()
                u = fu_var.get().strip()
                idx = df[df["Username"]==u].index
                if len(idx) == 0:
                    s2_lbl.config(text="User not found.", fg="#DC2626"); return
                df.loc[idx[0], "Password"] = np_
                df.to_excel(USERS_FILE(), index=False)
                messagebox.showinfo("✅ Password Reset",
                    f"Password changed successfully for '{u}'.\n\nYou can now login with your new password.",
                    parent=fw)
                fw.destroy()
            except Exception as ex:
                s2_lbl.config(text=f"Error: {ex}", fg="#DC2626")

        Button(step2, text="Set New Password",
               font=("Segoe UI",10,"bold"),
               bg="#16A34A", fg="#FFFFFF",
               relief=FLAT, cursor="hand2",
               activebackground="#15803D",
               pady=9, bd=0,
               command=save_new_password).pack(fill=X)

    bg_canvas.bind("<Configure>", _draw_bg)
    lw.after(50, _draw_bg)

    lw.bind("<Return>", lambda e: None)
    # Check for updates 3 seconds after login screen loads
    lw.after(3000, lambda: check_for_update_async(lw))
    lw.mainloop()



# ═══════════════════════════════════════════
#  ADMIN WINDOW
# ═══════════════════════════════════════════
def open_admin():
    root = Tk()
    root.title("GST — Admin Panel")
    root.geometry("1100x820")
    root.configure(bg=BG_PAGE)
    apply_modern_style()
    build_header(root, "ADMIN CONTROL PANEL")

    # ── Tab bar ──
    tab_bar = Frame(root, bg=BG_HEADER, pady=0)
    tab_bar.pack(fill=X)

    content = Frame(root, bg=BG_PAGE)
    content.pack(fill=BOTH, expand=True)

    t_log   = Frame(content, bg=BG_PAGE)
    t_users = Frame(content, bg=BG_PAGE)
    t_vis   = Frame(content, bg=BG_PAGE)
    t_bk    = Frame(content, bg=BG_PAGE)
    t_pur   = Frame(content, bg=BG_PAGE)
    t_off   = Frame(content, bg=BG_PAGE)
    t_cfg   = Frame(content, bg=BG_PAGE)
    t_vip   = Frame(content, bg=BG_PAGE)
    all_tabs = [t_log, t_users, t_vis, t_bk, t_pur, t_off, t_cfg, t_vip]
    tab_btns = []

    TAB_INFO = [("  📋  Staff Log  ",    t_log),
                ("  👥  Users  ",        t_users),
                ("  📊  Visitor Records  ", t_vis),
                ("  💾  Backup  ",        t_bk),
                ("  📋  Purpose List  ", t_pur),
                ("  🏢  Officers  ",     t_off),
                ("  ⚙️   Settings  ",     t_cfg),
                ("  ⭐  VIP System  ",   t_vip)]

    def switch_tab(i):
        for t in all_tabs: t.pack_forget()
        all_tabs[i].pack(fill=BOTH, expand=True)
        for j, b in enumerate(tab_btns):
            if j == i:
                b.config(bg="#3B82F6", fg=TEXT_WHITE,
                         font=("Segoe UI", 9, "bold"))
            else:
                b.config(bg=BG_HEADER, fg="#93C5FD",
                         font=("Segoe UI", 9))

    for i, (lbl, _) in enumerate(TAB_INFO):
        btn = Button(tab_bar, text=lbl, font=("Segoe UI", 9),
                     bg=BG_HEADER, fg="#93C5FD", relief=FLAT,
                     cursor="hand2", activebackground="#3B82F6",
                     activeforeground=TEXT_WHITE,
                     padx=10, pady=12, bd=0,
                     command=lambda i=i: switch_tab(i))
        btn.pack(side=LEFT)
        tab_btns.append(btn)
    Frame(tab_bar, bg="#3B82F6", height=3).pack(fill=X, side=BOTTOM)

    # ── TAB 1 — Staff Log ──
    _tab_title(t_log, "📋  Staff Login / Logout History")
    log_cols = ["Date","Username","Full Name","Role",
                "Login Time","Logout Time","Duration (mins)"]
    log_tree = build_treeview(t_log, log_cols)
    lbl_cnt  = Label(t_log, text="", font=("Segoe UI", 9), bg=BG_PAGE, fg=TEXT_MID)
    lbl_cnt.pack(anchor=W, padx=14)
    def refresh_log():
        for r in log_tree.get_children(): log_tree.delete(r)
        try:
            df = pd.read_excel(STAFF_LOG_FILE(), dtype=str)
            insert_tree_rows(log_tree, df)
            lbl_cnt.config(text=f"Total sessions: {len(df)}")
        except: pass
    def export_log():
        fn = f"GST_StaffLog_{datetime.now().strftime('%Y%m%d')}.xlsx"
        try:
            pd.read_excel(STAFF_LOG_FILE(), dtype=str).to_excel(fn, index=False)
            messagebox.showinfo("Exported", f"Saved:\n{os.path.abspath(fn)}")
        except Exception as ex: messagebox.showerror("Error", str(ex))
    br1 = Frame(t_log, bg=BG_PAGE, pady=6); br1.pack(fill=X, padx=14)
    make_button(br1, "📤  Export", BTN_SEARCH,
                command=export_log, padx=12, pady=7).pack(side=RIGHT, padx=4)
    make_button(br1, "🔄  Refresh", BG_HEADER,
                command=refresh_log, padx=12, pady=7).pack(side=RIGHT, padx=4)
    refresh_log()

    # ── TAB 2 — Manage Users ──
    _tab_title(t_users, "👥  Manage Staff Usernames & Passwords")
    ucols = ["Username","Full Name","Role","Mobile","Password"]
    utree = build_treeview(t_users, ucols, height=6)
    def refresh_users():
        for r in utree.get_children(): utree.delete(r)
        try:
            df = pd.read_excel(USERS_FILE(), dtype=str)
            for i, (_, r) in enumerate(df.iterrows()):
                tag = "odd" if i%2==0 else "even"
                mob = str(r.get("Mobile","")).strip() if "Mobile" in r.index else ""
                utree.insert("", END, values=[r["Username"],r["Full Name"],
                                               r["Role"],mob,"●●●●●●"], tags=(tag,))
        except: pass

    uform = make_card(t_users)
    uform.pack(fill=X, padx=14, pady=6)
    uf_inner = Frame(uform, bg=BG_CARD, padx=16, pady=12)
    uf_inner.pack(fill=X)
    uf_vars = {}
    for i, (lbl, key) in enumerate([("Username","un"),("Full Name","fn"),
                                      ("Mobile No","mb"),("Password","pw"),("Role","rl")]):
        Label(uf_inner, text=lbl, font=("Segoe UI", 9, "bold"),
              bg=BG_CARD, fg=TEXT_MID).grid(row=0, column=i*2, padx=(0,4), sticky=W)
        if key == "rl":
            v = StringVar(value="staff")
            cb = ttk.Combobox(uf_inner, textvariable=v, values=["staff","admin"],
                              font=("Segoe UI",9), width=10,
                              style="Modern.TCombobox", state="readonly")
            cb.grid(row=1, column=i*2, padx=(0,10), pady=(4,0), sticky=W)
            uf_vars[key] = v
        else:
            v = StringVar()
            e = make_entry(uf_inner, width=12, textvariable=v)
            e.grid(row=1, column=i*2, padx=(0,10), pady=(4,0), sticky=W, ipady=5)
            uf_vars[key] = v

    def add_user():
        un = uf_vars["un"].get().strip()
        fn = uf_vars["fn"].get().strip()
        pw = uf_vars["pw"].get().strip()
        rl = uf_vars["rl"].get().strip()
        mb = uf_vars["mb"].get().strip()
        if not un or not fn or not pw:
            messagebox.showwarning("Missing", "Fill Username, Full Name and Password."); return
        try:
            df = pd.read_excel(USERS_FILE(), dtype=str)
            if un in df["Username"].str.strip().values:
                messagebox.showwarning("Duplicate","Username already exists."); return
            if "Mobile" not in df.columns: df["Mobile"] = ""
            df = pd.concat([df, pd.DataFrame([{"Username":un,"Password":pw,
                                                "Full Name":fn,"Role":rl,"Mobile":mb}])],
                           ignore_index=True)
            df.to_excel(USERS_FILE(), index=False)
            refresh_users()
            for v in uf_vars.values(): v.set("")
            messagebox.showinfo("User Added", f"User '{un}' added.")
        except Exception as ex: messagebox.showerror("Error", str(ex))

    def delete_user():
        sel = utree.selection()
        if not sel: messagebox.showwarning("None","Select a user."); return
        un = utree.item(sel[0])["values"][0]
        if un == session["username"]:
            messagebox.showwarning("Blocked","Cannot delete currently logged-in user."); return
        if messagebox.askyesno("Confirm Delete", f"Delete user '{un}'?"):
            try:
                df = pd.read_excel(USERS_FILE(), dtype=str)
                df = df[df["Username"].str.strip() != un]
                df.to_excel(USERS_FILE(), index=False)
                refresh_users()
                messagebox.showinfo("Deleted", f"User '{un}' removed.")
            except Exception as ex: messagebox.showerror("Error", str(ex))

    def reset_pw():
        sel = utree.selection()
        if not sel: messagebox.showwarning("None","Select a user."); return
        un = utree.item(sel[0])["values"][0]
        pw = uf_vars["pw"].get().strip()
        if not pw: messagebox.showwarning("Missing","Enter new password in the Password field."); return
        try:
            df = pd.read_excel(USERS_FILE(), dtype=str)
            df.loc[df["Username"].str.strip()==un, "Password"] = pw
            df.to_excel(USERS_FILE(), index=False)
            refresh_users()
            messagebox.showinfo("Done", f"Password updated for '{un}'.")
        except Exception as ex: messagebox.showerror("Error", str(ex))

    ubr = Frame(t_users, bg=BG_PAGE, pady=6); ubr.pack(fill=X, padx=14)
    for txt, col, cmd in [
        ("➕  Add User",    BTN_IN,    add_user),
        ("🔑  Reset Pwd",   BTN_SEARCH, reset_pw),
        ("🗑  Delete User", BTN_CLEAR, delete_user),
        ("🔄  Refresh",     BG_HEADER, refresh_users),
    ]:
        make_button(ubr, txt, col, command=cmd, padx=10, pady=7).pack(side=LEFT, padx=4)
    refresh_users()

    # ── TAB 3 — Visitor Records ──
    _tab_title(t_vis, "📊  All Visitor Records")
    vis_cols = ["Group ID","Date","Arrival","Out","Visitor","Phone",
                "ID Cards","Total Members","Remaining","Company",
                "Purpose","GST No","Officer","Division","Block","Floor","Room No","Remarks"]
    vis_tree = build_treeview(t_vis, vis_cols, height=14)
    vcnt = Label(t_vis, text="", font=("Segoe UI",9), bg=BG_PAGE, fg=TEXT_MID)
    vcnt.pack(anchor=W, padx=14)
    def refresh_vis():
        for r in vis_tree.get_children(): vis_tree.delete(r)
        try:
            df = pd.read_excel(VISITORS_FILE(), dtype=str)
            insert_tree_rows(vis_tree, df[vis_cols])
            vcnt.config(text=f"Total records: {len(df)}")
        except: pass
    def export_vis():
        fn = f"GST_All_Visitors_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        try:
            pd.read_excel(VISITORS_FILE(), dtype=str).to_excel(fn, index=False)
            messagebox.showinfo("Exported", f"Saved:\n{os.path.abspath(fn)}")
        except Exception as ex: messagebox.showerror("Error", str(ex))
    vbr = Frame(t_vis, bg=BG_PAGE, pady=6); vbr.pack(fill=X, padx=14)
    make_button(vbr, "📤  Export All", BTN_SEARCH,
                command=export_vis, padx=12, pady=7).pack(side=RIGHT, padx=4)
    make_button(vbr, "🔄  Refresh", BG_HEADER,
                command=refresh_vis, padx=12, pady=7).pack(side=RIGHT, padx=4)
    refresh_vis()

    # ── TAB 4 — Backup ──
    _tab_title(t_bk, "💾  Backup & Restore")
    bk_card = make_card(t_bk)
    bk_card.pack(fill=X, padx=20, pady=12)
    bk_inner = Frame(bk_card, bg=BG_CARD, padx=24, pady=20)
    bk_inner.pack(fill=X)
    Label(bk_inner, text="Data Backup Options",
          font=("Segoe UI", 11, "bold"), bg=BG_CARD, fg=TEXT_DARK).pack(anchor=W)
    Label(bk_inner,
          text="All Excel data files are automatically backed up daily in GST_Backups\\ folder.",
          font=("Segoe UI", 9), bg=BG_CARD, fg=TEXT_MID).pack(anchor=W, pady=(4,16))
    Frame(bk_inner, bg=BORDER_CLR, height=1).pack(fill=X, pady=(0,12))
    bbr = Frame(bk_inner, bg=BG_CARD); bbr.pack(anchor=W)
    make_button(bbr, "💾  USB Backup", BTN_OUT,
                command=lambda: backup_to_usb(root),
                padx=14, pady=9).pack(side=LEFT, padx=(0,10))
    def manual_backup():
        auto_backup()
        messagebox.showinfo("Backup Done","Daily backup created in GST_Backups\\ folder.")
    make_button(bbr, "🔄  Run Manual Backup", BTN_SEARCH,
                command=manual_backup, padx=14, pady=9).pack(side=LEFT)

    # ── TAB 5 — Purpose List ──
    _tab_title(t_pur, "📋  Purpose Dropdown Manager")
    pur_card = make_card(t_pur)
    pur_card.pack(fill=BOTH, expand=True, padx=20, pady=12)
    pur_inner = Frame(pur_card, bg=BG_CARD, padx=20, pady=16)
    pur_inner.pack(fill=BOTH, expand=True)
    Label(pur_inner, text="Manage the 'Purpose of Visit' dropdown list",
          font=("Segoe UI", 10, "bold"), bg=BG_CARD, fg=TEXT_DARK).pack(anchor=W)
    Label(pur_inner, text="One purpose per line. Changes take effect immediately.",
          font=("Segoe UI", 9), bg=BG_CARD, fg=TEXT_MID).pack(anchor=W, pady=(4,10))

    pur_lb = Listbox(pur_inner, font=("Segoe UI",10), bg=FIELD_BG, fg=TEXT_DARK,
                     selectbackground=ACCENT_BLUE, selectforeground=TEXT_WHITE,
                     relief=FLAT, highlightthickness=1, highlightbackground=BORDER_CLR,
                     activestyle="none", height=10)
    pur_lb.pack(fill=BOTH, expand=True, pady=(0,8))

    pur_entry_var = StringVar()
    pur_row = Frame(pur_inner, bg=BG_CARD); pur_row.pack(fill=X)
    pur_e = make_entry(pur_row, width=36, textvariable=pur_entry_var)
    pur_e.pack(side=LEFT, ipady=5, padx=(0,8))

    def load_lb():
        pur_lb.delete(0, END)
        for p in load_purposes(): pur_lb.insert(END, p)
    def add_pur():
        p = pur_entry_var.get().strip()
        if not p: return
        items = list(pur_lb.get(0, END))
        if p not in items:
            items.append(p)
            save_purposes(items)
            load_lb()
            pur_entry_var.set("")
    def del_pur():
        sel = pur_lb.curselection()
        if not sel: return
        items = list(pur_lb.get(0, END))
        items.pop(sel[0])
        save_purposes(items)
        load_lb()
    make_button(pur_row, "➕ Add", BTN_IN, command=add_pur, padx=10, pady=7).pack(side=LEFT, padx=4)
    make_button(pur_row, "🗑 Delete", BTN_CLEAR, command=del_pur, padx=10, pady=7).pack(side=LEFT, padx=4)
    load_lb()

    # ── TAB 6 — Officers ──
    _tab_title(t_off, "🏢  Officer Management — Import / Export / Edit")

    off_body = Frame(t_off, bg=BG_PAGE)
    off_body.pack(fill=BOTH, expand=True, padx=14, pady=8)

    # Top info + buttons
    off_top = make_card(off_body)
    off_top.pack(fill=X, pady=(0,8))
    off_top_in = Frame(off_top, bg=BG_CARD, padx=16, pady=12); off_top_in.pack(fill=X)
    Label(off_top_in,
          text="Upload your officer list Excel file. Receptionist types an officer name/code and fields auto-fill.",
          font=("Segoe UI",9), bg=BG_CARD, fg=TEXT_MID).pack(anchor=W, pady=(0,10))

    # Template download hint
    Label(off_top_in,
          text="Excel columns required:  Officer Code  |  Officer Name  |  Designation  |  Division  |  Room No  |  Block  |  Floor",
          font=("Segoe UI",8,"italic"), bg=BG_CARD, fg=TEXT_LIGHT).pack(anchor=W, pady=(0,8))

    off_btn_row = Frame(off_top_in, bg=BG_CARD); off_btn_row.pack(anchor=W)

    off_cnt_lbl = Label(off_body, text="", font=("Segoe UI",9),
                        bg=BG_PAGE, fg=TEXT_MID)
    off_cnt_lbl.pack(anchor=W, padx=4, pady=(0,4))

    # Treeview for officer list
    off_tree = build_treeview(off_body, _OFFICER_COLS, height=14)

    def refresh_officers():
        for r in off_tree.get_children(): off_tree.delete(r)
        recs = load_officers()
        for i, o in enumerate(recs):
            tag = "odd" if i%2==0 else "even"
            off_tree.insert("", END, values=[o.get(c,"") for c in _OFFICER_COLS], tags=(tag,))
        off_cnt_lbl.config(text=f"Total officers: {len(recs)}")

    def import_officers():
        from tkinter import filedialog
        fp = filedialog.askopenfilename(
            title="Select Officer Excel File",
            filetypes=[("Excel files","*.xlsx *.xls"),("All files","*.*")])
        if not fp: return
        try:
            df = pd.read_excel(fp, dtype=str).fillna("")
            # flexible column matching — case-insensitive
            df.columns = [c.strip() for c in df.columns]
            col_map = {}
            for want in _OFFICER_COLS:
                for have in df.columns:
                    if want.lower() == have.lower():
                        col_map[want] = have
                        break
            records = []
            for _, row in df.iterrows():
                rec = {c: str(row.get(col_map.get(c,""), "")).strip()
                       for c in _OFFICER_COLS}
                records.append(rec)
            save_officers(records)
            refresh_officers()
            messagebox.showinfo("Import Successful", "Imported " + str(len(records)) + " officer records.\n\nSource: " + os.path.basename(fp))
        except Exception as ex:
            messagebox.showerror("Import Error", str(ex))

    def export_officers():
        recs = load_officers()
        if not recs:
            messagebox.showwarning("Empty","No officer records to export."); return
        fn = f"GST_Officers_Export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        fp = os.path.join(load_data_path(), fn)
        save_officers(recs)
        import shutil as _sh
        _sh.copy2(OFFICERS_FILE(), fp)
        messagebox.showinfo("Exported", "Saved to:\n" + str(fp))

    def download_template():
        fn = "GST_Officers_Template.xlsx"
        fp = os.path.join(load_data_path(), fn)
        sample = [
            {"Officer Code":"CTO11","Officer Name":"Mr. K. Sharma",
             "Designation":"Chief Technical Officer","Division":"Enforcement",
             "Room No":"305","Block":"A Block","Floor":"1st Floor"},
            {"Officer Code":"DCO05","Officer Name":"Ms. Priya Nair",
             "Designation":"Deputy Commissioner","Division":"Audit",
             "Room No":"201","Block":"B Block","Floor":"2nd Floor"},
        ]
        pd.DataFrame(sample, columns=_OFFICER_COLS).to_excel(fp, index=False)
        messagebox.showinfo("Template Ready",
            "Template saved to:\n" + str(fp) + "\n\n"
            "Fill in your officer data and import it.")
    def delete_officer():
        sel = off_tree.selection()
        if not sel:
            messagebox.showwarning("None","Select an officer to delete."); return
        vals = off_tree.item(sel[0])["values"]
        code = str(vals[0]) if vals else ""
        if not messagebox.askyesno("Confirm Delete",
                f"Delete officer: {vals[1]} ({code})?"):
            return
        recs = [o for o in load_officers()
                if o.get("Officer Code","").strip() != code]
        save_officers(recs)
        refresh_officers()

    def clear_all_officers():
        if not messagebox.askyesno("Confirm Clear All", "Delete ALL officer records?\n\nThis cannot be undone."):
            return
        save_officers([])
        refresh_officers()

    make_button(off_btn_row, "📥  Import Excel", BTN_IN,
                command=import_officers, padx=12, pady=8).pack(side=LEFT, padx=(0,6))
    make_button(off_btn_row, "📤  Export Excel", BTN_SEARCH,
                command=export_officers, padx=12, pady=8).pack(side=LEFT, padx=6)
    make_button(off_btn_row, "📋  Download Template", BG_HEADER,
                command=download_template, padx=12, pady=8).pack(side=LEFT, padx=6)
    make_button(off_btn_row, "🗑  Delete Selected", BTN_CLEAR,
                command=delete_officer, padx=12, pady=8).pack(side=LEFT, padx=6)
    make_button(off_btn_row, "❌  Clear All", ACCENT_RED,
                command=clear_all_officers, padx=12, pady=8).pack(side=LEFT, padx=6)

    refresh_officers()

    # ── TAB 7 — Settings ──
    _tab_title(t_cfg, "⚙️   System Settings  —  Admin Only")
    cfg_scroll = Frame(t_cfg, bg=BG_PAGE)
    cfg_scroll.pack(fill=BOTH, expand=True, padx=20, pady=10)

    _cur = load_data_path()
    _cur_low = _cur.lower()
    _is_risky = ("\\temp" in _cur_low or "/temp" in _cur_low or
                 "appdata\\local\\temp" in _cur_low)
    if _is_risky:
        warn_card = make_card(cfg_scroll)
        warn_card.configure(highlightbackground="#FCA5A5", bg="#FEF2F2")
        warn_card.pack(fill=X, pady=(0,10))
        wi = Frame(warn_card, bg="#FEF2F2", padx=16, pady=12); wi.pack(fill=X)
        Label(wi, text="⚠️  WARNING — Data is in a TEMP folder!",
              font=("Segoe UI",11,"bold"), bg="#FEF2F2", fg="#B91C1C").pack(anchor=W)
        Label(wi, text="Temp folders can be deleted automatically by Windows.\n"
                       "Please choose a safe folder below.",
              font=("Segoe UI",9), bg="#FEF2F2", fg="#7F1D1D").pack(anchor=W, pady=(4,0))

    path_card = make_card(cfg_scroll)
    path_card.pack(fill=X, pady=(0,12))
    pi = Frame(path_card, bg=BG_CARD, padx=20, pady=18); pi.pack(fill=X)

    Label(pi, text="📁  Data Save Location",
          font=("Segoe UI",11,"bold"), bg=BG_CARD, fg=TEXT_DARK).pack(anchor=W)
    Label(pi, text="All Excel files (Visitors, Staff Log, Users) and photos are saved here.",
          font=("Segoe UI",9), bg=BG_CARD, fg=TEXT_MID).pack(anchor=W, pady=(2,8))

    cur_path_var = StringVar(value=_cur)
    cp_box = Frame(pi, bg=STATUS_BG, bd=0, relief=FLAT,
                   highlightthickness=1, highlightbackground="#BFDBFE")
    cp_box.pack(fill=X, pady=(0,8))
    cpi = Frame(cp_box, bg=STATUS_BG, padx=12, pady=10); cpi.pack(fill=X)
    Label(cpi, text="Current save location:", font=("Segoe UI",9,"bold"),
          bg=STATUS_BG, fg=TEXT_MID).pack(anchor=W)
    cur_path_lbl = Label(cpi, textvariable=cur_path_var,
                         font=("Segoe UI",10,"bold"), bg=STATUS_BG,
                         fg=ACCENT_GREEN, wraplength=600, justify=LEFT)
    cur_path_lbl.pack(anchor=W, pady=(4,0))

    def open_data_folder():
        p = load_data_path()
        if os.path.exists(p): os.startfile(p)
        else: messagebox.showinfo("Not Found", f"Folder not found:\n{p}")
    make_button(cpi, "📂  Open in Explorer", BG_HEADER,
                command=open_data_folder, padx=12, pady=7).pack(anchor=W, pady=(8,0))

    files_lbl = Label(pi, text="", font=("Segoe UI",8),
                      bg=BG_CARD, fg=TEXT_MID, justify=LEFT)
    files_lbl.pack(anchor=W, pady=(6,0))

    def refresh_path_display():
        p = load_data_path()
        cur_path_var.set(p)
        names = [_VISITORS_FNAME, _STAFFLOG_FNAME, _USERS_FNAME, _PHOTOS_DIRNAME]
        lines = []
        for n in names:
            fp = os.path.join(p, n)
            exists = "✅" if os.path.exists(fp) else "⬜  (created on first save)"
            lines.append(f"  {exists}  {fp}")
        files_lbl.config(text="\n".join(lines))
    refresh_path_display()

    Frame(pi, bg=BORDER_CLR, height=1).pack(fill=X, pady=(14,10))
    Label(pi, text="Change save location:",
          font=("Segoe UI",10,"bold"), bg=BG_CARD, fg=TEXT_DARK).pack(anchor=W, pady=(0,4))
    Label(pi, text="Recommended:  C:\\Users\\YourName\\Documents\\GST_Data",
          font=("Segoe UI",8,"italic"), bg=BG_CARD, fg=TEXT_LIGHT).pack(anchor=W, pady=(0,6))

    np_row = Frame(pi, bg=BG_CARD); np_row.pack(fill=X)
    new_path_var = StringVar()
    new_path_entry = make_entry(np_row, width=52, textvariable=new_path_var)
    new_path_entry.pack(side=LEFT, padx=(0,8), ipady=6)

    def browse_folder():
        from tkinter import filedialog
        chosen = filedialog.askdirectory(title="Select folder to save GST data files",
                                         initialdir=load_data_path())
        if chosen: new_path_var.set(chosen.replace("/","\\"))
    make_button(np_row, "📂  Browse", BG_HEADER,
                command=browse_folder, padx=10, pady=7).pack(side=LEFT, padx=4)

    def apply_new_path():
        np = new_path_var.get().strip().replace("/","\\")
        if not np:
            messagebox.showwarning("Empty","Enter or browse a folder path first.", parent=root); return
        np_low = np.lower()
        if "\\temp\\" in np_low or np_low.endswith("\\temp") or "appdata\\local\\temp" in np_low:
            messagebox.showerror("❌ Unsafe Folder",
                "You cannot save data to a Temp folder.\n\n"
                "Temp folders are deleted by Windows automatically.\n"
                "Please choose Documents, Desktop, or D:\\ drive.", parent=root); return
        if not os.path.isdir(np):
            create = messagebox.askyesno("Folder Not Found",
                f"Folder does not exist:\n{np}\n\nCreate it now?", parent=root)
            if not create: return
            try: os.makedirs(np, exist_ok=True)
            except Exception as ex:
                messagebox.showerror("Error", f"Could not create folder:\n{ex}", parent=root); return
        migrate = messagebox.askyesno("Copy Existing Data?",
            f"New save location:\n{np}\n\n"
            f"Copy your existing data files to the new location?\n\n"
            f"Yes = Copy all files (RECOMMENDED)\n"
            f"No  = Start fresh in new folder", parent=root)
        if migrate:
            moved = []
            old_path = load_data_path()
            for fname in [_VISITORS_FNAME, _STAFFLOG_FNAME, _USERS_FNAME]:
                src = os.path.join(old_path, fname)
                dst = os.path.join(np, fname)
                if os.path.exists(src):
                    shutil.copy2(src, dst)
                    moved.append(f"✅  {fname}")
            src_photos = os.path.join(old_path, _PHOTOS_DIRNAME)
            dst_photos = os.path.join(np, _PHOTOS_DIRNAME)
            if os.path.exists(src_photos):
                if os.path.exists(dst_photos): shutil.rmtree(dst_photos)
                shutil.copytree(src_photos, dst_photos)
                moved.append(f"✅  {_PHOTOS_DIRNAME}\\")
            if moved:
                messagebox.showinfo("✅ Files Copied",
                    "Successfully copied to new location:\n\n" + "\n".join(moved), parent=root)
        save_data_path(np)
        new_path_var.set("")
        refresh_path_display()
        messagebox.showinfo("✅ Path Saved",
            f"Data will now be saved to:\n\n{np}\n\n"
            f"This setting is permanent.", parent=root)

    make_button(pi, "✅  Apply & Save New Path", BTN_IN,
                command=apply_new_path, padx=14, pady=9).pack(anchor=W, pady=(10,0))

    Frame(cfg_scroll, bg=BORDER_CLR, height=1).pack(fill=X, pady=14)
    rst_card = make_card(cfg_scroll)
    rst_card.pack(fill=X)
    ri = Frame(rst_card, bg="#FFFBEB", padx=20, pady=14); ri.pack(fill=X)
    Label(ri, text="🔄  Quick Reset — Save to Desktop",
          font=("Segoe UI",10,"bold"), bg="#FFFBEB", fg=BTN_OUT).pack(anchor=W)
    _desktop = os.path.join(os.path.expanduser("~"), "Desktop", "GST_Data")
    Label(ri, text=f"Resets save location to: {_desktop}",
          font=("Segoe UI",9,"italic"), bg="#FFFBEB", fg=TEXT_MID).pack(anchor=W, pady=4)
    def reset_to_desktop():
        target = _desktop
        if messagebox.askyesno("Reset to Desktop",
            f"Save all data to:\n{target}\n\nExisting files will be copied. Continue?", parent=root):
            try: os.makedirs(target, exist_ok=True)
            except Exception as ex:
                messagebox.showerror("Error", str(ex), parent=root); return
            old = load_data_path()
            for fname in [_VISITORS_FNAME, _STAFFLOG_FNAME, _USERS_FNAME]:
                src = os.path.join(old, fname); dst = os.path.join(target, fname)
                if os.path.exists(src) and not os.path.exists(dst): shutil.copy2(src, dst)
            src_ph = os.path.join(old, _PHOTOS_DIRNAME)
            dst_ph = os.path.join(target, _PHOTOS_DIRNAME)
            if os.path.exists(src_ph) and not os.path.exists(dst_ph): shutil.copytree(src_ph, dst_ph)
            save_data_path(target)
            refresh_path_display()
            messagebox.showinfo("✅ Done", f"Data location set to:\n{target}", parent=root)
    make_button(ri, "🔄  Reset to Desktop\\GST_Data", BTN_OUT,
                command=reset_to_desktop, padx=12, pady=7).pack(anchor=W, pady=6)

    # ── TAB: VIP System Settings ─────────────────────────────────
    _tab_title(t_vip, "⭐  VIP System Settings")

    vip_body = Frame(t_vip, bg=BG_PAGE, padx=24, pady=20)
    vip_body.pack(fill=BOTH, expand=True)

    # Info card
    info_box = Frame(vip_body, bg="#FFFBEB",
                     highlightthickness=1, highlightbackground="#FCD34D")
    info_box.pack(fill=X, pady=(0,20))
    Label(info_box, text="ℹ️  What is the VIP System?",
          font=("Segoe UI",10,"bold"),
          bg="#FFFBEB", fg="#92400E",
          padx=16, pady=8).pack(anchor=W)
    Label(info_box,
          text=(
              "When a visitor comes 3 or more times, the system offers to generate\n"
              "a VIP code for them (e.g. B7766, CTO7).\n\n"
              "With a VIP code:\n"
              "  • Receptionist types code → all details load instantly\n"
              "  • Faster entry for frequent visitors\n"
              "  • VIP badge shown in returning visitor banner\n\n"
              "Turn OFF if your office does not use VIP codes."
          ),
          font=("Segoe UI",9),
          bg="#FFFBEB", fg="#92400E",
          padx=16, pady=6, justify=LEFT).pack(anchor=W)

    # Current status display
    _vip_currently = load_vip_enabled()
    status_lbl_var = StringVar(
        value="✅  VIP System is currently: ON" if _vip_currently
              else "❌  VIP System is currently: OFF")
    status_color = "#16A34A" if _vip_currently else "#DC2626"

    status_card = Frame(vip_body, bg=BG_CARD,
                        highlightthickness=1, highlightbackground=BORDER_CLR)
    status_card.pack(fill=X, pady=(0,20))
    status_inner = Frame(status_card, bg=BG_CARD, padx=20, pady=16)
    status_inner.pack(fill=X)

    status_display = Label(status_inner, textvariable=status_lbl_var,
                           font=("Segoe UI",13,"bold"),
                           bg=BG_CARD, fg=status_color)
    status_display.pack(anchor=W)
    Label(status_inner,
          text="Change takes effect immediately for all receptionists.",
          font=("Segoe UI",8,"italic"),
          bg=BG_CARD, fg=TEXT_LIGHT).pack(anchor=W, pady=(4,0))

    # Toggle buttons
    btn_row_vip = Frame(vip_body, bg=BG_PAGE)
    btn_row_vip.pack(anchor=W, pady=(0,20))

    def _set_vip(enabled):
        if save_vip_enabled(enabled):
            if enabled:
                status_lbl_var.set("✅  VIP System is currently: ON")
                status_display.config(fg="#16A34A")
                messagebox.showinfo(
                    "✅  VIP System — ON",
                    "VIP System is now ENABLED.\n\n"
                    "Receptionist will see VIP code prompt\n"
                    "when a visitor comes 3+ times.\n\n"
                    "VIP entry field is active on reception screen.",
                    parent=root)
            else:
                status_lbl_var.set("❌  VIP System is currently: OFF")
                status_display.config(fg="#DC2626")
                messagebox.showinfo(
                    "❌  VIP System — OFF",
                    "VIP System is now DISABLED.\n\n"
                    "No VIP code prompts will appear.\n"
                    "VIP entry field is hidden on reception screen.\n\n"
                    "Existing VIP codes in Excel are kept safely.",
                    parent=root)
        else:
            messagebox.showerror("Error",
                "Could not save setting. Check file permissions.",
                parent=root)

    make_button(btn_row_vip, "✅  Turn VIP ON", BTN_IN,
                command=lambda: _set_vip(True),
                padx=18, pady=10).pack(side=LEFT, padx=(0,12))

    make_button(btn_row_vip, "❌  Turn VIP OFF", ACCENT_RED,
                command=lambda: _set_vip(False),
                padx=18, pady=10).pack(side=LEFT)

    # What changes when OFF
    off_box = Frame(vip_body, bg="#FEF2F2",
                    highlightthickness=1, highlightbackground="#FECACA")
    off_box.pack(fill=X)
    Label(off_box, text="When VIP is OFF:",
          font=("Segoe UI",9,"bold"),
          bg="#FEF2F2", fg="#DC2626",
          padx=16, pady=8).pack(anchor=W)
    Label(off_box,
          text=(
              "  • VIP code entry field hidden from receptionist\n"
              "  • No VIP code popup on 3rd visit\n"
              "  • VIP badge not shown in returning visitor banner\n"
              "  • All existing data and Excel records stay safe"
          ),
          font=("Segoe UI",9),
          bg="#FEF2F2", fg="#B91C1C",
          padx=16, pady=6, justify=LEFT).pack(anchor=W)

    build_bottom_bar(root, lambda: do_logout(root))
    switch_tab(0)
    root.protocol("WM_DELETE_WINDOW", lambda: do_logout(root))
    root.mainloop()


def _tab_title(parent, text):
    """Consistent tab section title bar."""
    bar = Frame(parent, bg=BG_CARD, bd=0, relief=FLAT,
                highlightthickness=1, highlightbackground=BORDER_CLR)
    bar.pack(fill=X, padx=14, pady=(12,0))
    Label(bar, text=text, font=("Segoe UI", 11, "bold"),
          bg=BG_CARD, fg=TEXT_DARK, pady=12, padx=16).pack(side=LEFT)



# ═══════════════════════════════════════════
#  RECEPTION WINDOW — matches Image 1 exactly
#  Dark navy sidebar LEFT + white form RIGHT
# ═══════════════════════════════════════════
def open_reception():
    # ── Colors — warm professional MNC style ──
    SB_BG    = "#1B2E4B"   # sidebar dark navy
    SB_ACT   = "#2563EB"   # active menu blue
    SB_TEXT  = "#CBD5E1"   # sidebar normal text
    SB_HEAD  = "#0F1F35"   # sidebar top darker
    TOP_BG   = "#F8F9FA"   # warm off-white stats bar
    FRM_BG   = "#F0F2F5"   # warm light gray background
    CARD_W   = "#FFFFFF"   # form card white

    root = Tk()
    root.title("GST — Visitor Management System")
    root.state("zoomed")
    root.configure(bg=FRM_BG)
    apply_modern_style()

    # ══════════════════════════════════════
    #  OUTER LAYOUT: sidebar LEFT + main RIGHT
    # ══════════════════════════════════════
    outer = Frame(root, bg=FRM_BG)
    outer.pack(fill=BOTH, expand=True)

    # ── LEFT SIDEBAR ──
    sidebar = Frame(outer, bg=SB_BG, width=220)
    sidebar.pack(side=LEFT, fill=Y)
    sidebar.pack_propagate(False)

    # Sidebar top: logo + title
    sb_top = Frame(sidebar, bg=SB_HEAD, pady=20)
    sb_top.pack(fill=X)
    logo_cv = Canvas(sb_top, width=48, height=48, bg=SB_HEAD,
                     highlightthickness=0)
    logo_cv.pack()
    logo_cv.create_oval(4, 4, 44, 44, fill="#2D5282", outline="#4A7CBC", width=1.5)
    logo_cv.create_text(24, 24, text="🏛", font=("Segoe UI Emoji", 22), fill="#F6D860")

    Label(sb_top, text="VISITOR",
          font=("Segoe UI", 9, "bold"), bg=SB_HEAD, fg="#FFFFFF").pack()
    Label(sb_top, text="MANAGEMENT",
          font=("Segoe UI", 9, "bold"), bg=SB_HEAD, fg="#FFFFFF").pack()
    Label(sb_top, text="SYSTEM",
          font=("Segoe UI", 9, "bold"), bg=SB_HEAD, fg="#FFFFFF").pack()
    Frame(sb_top, bg="#2563EB", height=2).pack(fill=X, pady=(10,0))
    Label(sb_top, text="Goods and Services Tax\nDepartment",
          font=("Segoe UI", 7), bg=SB_HEAD, fg="#64748B",
          justify=CENTER).pack(pady=(4,0))

    # Menu items
    sb_menu = Frame(sidebar, bg=SB_BG)
    sb_menu.pack(fill=X, pady=(8,0))

    def _menu_item(parent, icon, text, active=False, command=None):
        bg = SB_ACT if active else SB_BG
        f = Frame(parent, bg=bg, cursor="hand2")
        f.pack(fill=X)
        inner = Frame(f, bg=bg, pady=12, padx=16)
        inner.pack(fill=X)
        Label(inner, text=icon, font=("Segoe UI Emoji", 12),
              bg=bg, fg="#FFFFFF").pack(side=LEFT, padx=(0,10))
        Label(inner, text=text, font=("Segoe UI", 10, "bold" if active else "normal"),
              bg=bg, fg="#FFFFFF" if active else SB_TEXT).pack(side=LEFT)
        if command:
            f.bind("<Button-1>", lambda e: command())
            for w in f.winfo_children():
                w.bind("<Button-1>", lambda e: command())
                for ww in w.winfo_children():
                    ww.bind("<Button-1>", lambda e: command())
        # Hover
        def _enter(e, _f=f, _bg=bg):
            if not active: _f.config(bg="#243552")
            for w in _f.winfo_children():
                w.config(bg="#243552" if not active else SB_ACT)
        def _leave(e, _f=f, _bg=bg):
            _f.config(bg=_bg)
            for w in _f.winfo_children():
                w.config(bg=_bg)
        f.bind("<Enter>", _enter); f.bind("<Leave>", _leave)
        return f

    _menu_item(sb_menu, "⊞", "Dashboard",    active=True)
    _menu_item(sb_menu, "👥", "Visitors",     command=lambda: show_today_detail())
    _menu_item(sb_menu, "📊", "Daily Report", command=lambda: daily_report())
    _menu_item(sb_menu, "⚙",  "Settings",    command=lambda: _show_reception_settings())
    _menu_item(sb_menu, "⬆",  "Check Update", command=lambda: check_for_update_async(root, manual=True))

    # ── Logout button — prominent in sidebar ──
    Frame(sb_menu, bg="#243552", height=1).pack(fill=X, pady=(16,0))
    logout_f = Frame(sb_menu, bg="#DC2626", cursor="hand2")
    logout_f.pack(fill=X)
    logout_in = Frame(logout_f, bg="#DC2626", pady=13, padx=16)
    logout_in.pack(fill=X)
    Label(logout_in, text="🔒", font=("Segoe UI Emoji",12),
          bg="#DC2626", fg="#FFFFFF").pack(side=LEFT, padx=(0,10))
    Label(logout_in, text="Logout",
          font=("Segoe UI",10,"bold"), bg="#DC2626", fg="#FFFFFF").pack(side=LEFT)
    def _do_logout_btn():
        do_logout(root)
    logout_f.bind("<Button-1>", lambda e: _do_logout_btn())
    logout_in.bind("<Button-1>", lambda e: _do_logout_btn())
    for w in logout_in.winfo_children():
        w.bind("<Button-1>", lambda e: _do_logout_btn())
    def _lo_enter(e):
        logout_f.config(bg="#B91C1C"); logout_in.config(bg="#B91C1C")
        for w in logout_in.winfo_children(): w.config(bg="#B91C1C")
    def _lo_leave(e):
        logout_f.config(bg="#DC2626"); logout_in.config(bg="#DC2626")
        for w in logout_in.winfo_children(): w.config(bg="#DC2626")
    logout_f.bind("<Enter>", _lo_enter); logout_f.bind("<Leave>", _lo_leave)

    # Sidebar bottom: date/time + user
    sb_bot = Frame(sidebar, bg=SB_BG)
    sb_bot.pack(side=BOTTOM, fill=X, pady=8)
    Frame(sb_bot, bg="#243552", height=1).pack(fill=X, pady=(0,8))

    sb_date_v = StringVar(); sb_time_v = StringVar()
    def _sb_clk():
        n = datetime.now()
        sb_date_v.set(n.strftime("%d %B %Y"))
        sb_time_v.set(n.strftime("%H:%M:%S"))
        root.after(1000, _sb_clk)
    _sb_clk()

    date_row = Frame(sb_bot, bg=SB_BG, padx=14)
    date_row.pack(fill=X)
    Label(date_row, text="📅", font=("Segoe UI Emoji", 10),
          bg=SB_BG, fg=SB_TEXT).pack(side=LEFT, padx=(0,6))
    Label(date_row, textvariable=sb_date_v,
          font=("Segoe UI", 9), bg=SB_BG, fg=SB_TEXT).pack(side=LEFT)

    time_row = Frame(sb_bot, bg=SB_BG, padx=14)
    time_row.pack(fill=X, pady=(4,8))
    Label(time_row, text="🕐", font=("Segoe UI Emoji", 10),
          bg=SB_BG, fg=SB_TEXT).pack(side=LEFT, padx=(0,6))
    Label(time_row, textvariable=sb_time_v,
          font=("Segoe UI", 9), bg=SB_BG, fg=SB_TEXT).pack(side=LEFT)

    Frame(sb_bot, bg="#243552", height=1).pack(fill=X, pady=(0,8))
    user_row = Frame(sb_bot, bg=SB_BG, padx=14, pady=6)
    user_row.pack(fill=X)
    Label(user_row, text="👤", font=("Segoe UI Emoji", 14),
          bg="#243552", fg="#FFFFFF", padx=6, pady=4).pack(side=LEFT)
    uf2 = Frame(user_row, bg=SB_BG); uf2.pack(side=LEFT, padx=8)
    Label(uf2, text=session.get("full_name","User").upper(),
          font=("Segoe UI", 8, "bold"), bg=SB_BG, fg="#FFFFFF").pack(anchor=W)
    Label(uf2, text=session.get("role","STAFF").upper(),
          font=("Segoe UI", 7), bg=SB_BG, fg="#64748B").pack(anchor=W)

    # ── RIGHT MAIN AREA ──
    main = Frame(outer, bg=FRM_BG)
    main.pack(side=LEFT, fill=BOTH, expand=True)

    # ── TOP STATS BAR ──
    stats_bar = Frame(main, bg=TOP_BG,
                      highlightthickness=1, highlightbackground=BORDER_CLR)
    stats_bar.pack(fill=X)
    sb_inner = Frame(stats_bar, bg=TOP_BG, pady=10, padx=16)
    sb_inner.pack(fill=X)

    # Title left
    title_f = Frame(sb_inner, bg=TOP_BG); title_f.pack(side=LEFT)
    Label(title_f, text="Visitor Registration",
          font=("Segoe UI", 14, "bold"), bg=TOP_BG, fg=TEXT_DARK).pack(anchor=W)
    Label(title_f, text="Dashboard / Visitor Management",
          font=("Segoe UI", 8), bg=TOP_BG, fg=ACCENT_BLUE).pack(anchor=W)

    # Stats pills right
    cv_s  = StringVar(value="Today: 0")
    iv_s  = StringVar(value="Inside: 0")
    cam_s = StringVar(value="📷  Webcam: Ready")

    def _spill(par, var, bg, fg, icon="", cmd=None):
        f = Frame(par, bg=bg, bd=0, relief=FLAT,
                  highlightthickness=1, highlightbackground=BORDER_CLR,
                  cursor="hand2" if cmd else "")
        f.pack(side=LEFT, padx=(0,8))
        inner = Frame(f, bg=bg, padx=10, pady=5)
        inner.pack()
        if icon:
            Label(inner, text=icon, font=("Segoe UI Emoji", 9),
                  bg=bg, fg=fg).pack(side=LEFT, padx=(0,4))
        Label(inner, textvariable=var, font=("Segoe UI", 9, "bold"),
              bg=bg, fg=fg).pack(side=LEFT)
        if cmd:
            for w in [f, inner] + list(inner.winfo_children()):
                try: w.bind("<Button-1>", lambda e, c=cmd: c())
                except: pass

    pills_f = Frame(sb_inner, bg=TOP_BG); pills_f.pack(side=RIGHT)
    scan_s = StringVar(value="🔌  Scanner: Checking...")



    # ── HIGH PERFORMANCE CACHE ─────────────────────────────────
    # Reads Excel only when data changes — not on every action
    # openpyxl engine is 3x faster than default for reading
    _df_cache       = [None]
    _df_cache_time  = [0.0]
    _CACHE_TTL      = 8   # seconds before auto-refresh

    def get_cached_df(force=False):
        """
        Return cached visitors DataFrame.
        Re-reads from Excel only if:
          - cache is empty
          - cache is older than _CACHE_TTL seconds
          - force=True (called after save/delete)
        """
        import time
        now = time.time()
        if (force or
                _df_cache[0] is None or
                (now - _df_cache_time[0]) > _CACHE_TTL):
            try:
                _df_cache[0] = pd.read_excel(
                    VISITORS_FILE(),
                    dtype=str,
                    engine="openpyxl"
                ).fillna("")
                _df_cache_time[0] = now
            except Exception as _ce:
                print(f"Cache read error: {_ce}")
                if _df_cache[0] is None:
                    return pd.DataFrame()
        return _df_cache[0].copy()

    def invalidate_cache():
        """Force fresh read on next access — call after every save."""
        _df_cache[0] = None
        _df_cache_time[0] = 0.0

    def _show_reception_settings():
        """Reception Settings — 3 options: Change Password, About, Check Update."""
        top = Toplevel(root)
        top.title("Settings")
        top.geometry("440x520")
        top.resizable(False, False)
        top.configure(bg="#FFFFFF")
        top.transient(root)
        top.grab_set()
        top.update_idletasks()
        sw = top.winfo_screenwidth()
        sh = top.winfo_screenheight()
        top.geometry(f"440x520+{(sw-440)//2}+{(sh-520)//2}")

        # Header
        hdr = Frame(top, bg="#1B2E4B"); hdr.pack(fill=X)
        Label(hdr, text="⚙  Settings",
              font=("Segoe UI",13,"bold"),
              bg="#1B2E4B", fg="#C8A84B", pady=12).pack()

        body = Frame(top, bg="#FFFFFF", padx=22, pady=16)
        body.pack(fill=BOTH, expand=True)

        # ── OPTION 1: Change My Password ─────────────────────────
        pw_card = Frame(body, bg="#F8FAFF",
                        highlightthickness=1, highlightbackground="#E2E8F0")
        pw_card.pack(fill=X, pady=(0,14))
        pw_in = Frame(pw_card, bg="#F8FAFF", padx=16, pady=14)
        pw_in.pack(fill=X)

        Label(pw_in, text="🔑  Change My Password",
              font=("Segoe UI",10,"bold"),
              bg="#F8FAFF", fg="#1E293B").pack(anchor=W)
        Label(pw_in, text="Change your login password securely.",
              font=("Segoe UI",8,"italic"),
              bg="#F8FAFF", fg="#94A3B8").pack(anchor=W, pady=(2,10))

        Label(pw_in, text="Current Password:",
              font=("Segoe UI",9), bg="#F8FAFF", fg="#475569").pack(anchor=W)
        cur_pw = Entry(pw_in, show="●", font=("Segoe UI",10),
                       relief=FLAT, bg="#FFFFFF",
                       highlightthickness=1, highlightbackground="#CBD5E1",
                       highlightcolor="#2563EB")
        cur_pw.pack(fill=X, ipady=6, pady=(2,8))

        Label(pw_in, text="New Password:",
              font=("Segoe UI",9), bg="#F8FAFF", fg="#475569").pack(anchor=W)
        new_pw = Entry(pw_in, show="●", font=("Segoe UI",10),
                       relief=FLAT, bg="#FFFFFF",
                       highlightthickness=1, highlightbackground="#CBD5E1",
                       highlightcolor="#2563EB")
        new_pw.pack(fill=X, ipady=6, pady=(2,8))

        Label(pw_in, text="Confirm New Password:",
              font=("Segoe UI",9), bg="#F8FAFF", fg="#475569").pack(anchor=W)
        con_pw = Entry(pw_in, show="●", font=("Segoe UI",10),
                       relief=FLAT, bg="#FFFFFF",
                       highlightthickness=1, highlightbackground="#CBD5E1",
                       highlightcolor="#2563EB")
        con_pw.pack(fill=X, ipady=6, pady=(2,8))

        pw_msg = Label(pw_in, text="", font=("Segoe UI",8,"bold"),
                       bg="#F8FAFF", fg="#DC2626")
        pw_msg.pack(anchor=W)

        def _change_password():
            cur = cur_pw.get().strip()
            nw  = new_pw.get().strip()
            cn  = con_pw.get().strip()
            if not cur or not nw or not cn:
                pw_msg.config(text="❌  All fields are required.", fg="#DC2626")
                return
            if nw != cn:
                pw_msg.config(text="❌  New passwords do not match.", fg="#DC2626")
                return
            if len(nw) < 3:
                pw_msg.config(text="❌  Password too short (min 3 chars).", fg="#DC2626")
                return
            try:
                uname = session.get("username","")
                df = pd.read_excel(USERS_FILE(), dtype=str)
                df["Username"] = df["Username"].astype(str).str.strip()
                df["Password"] = df["Password"].astype(str).str.strip()
                match = df[(df["Username"]==uname) & (df["Password"]==cur)]
                if match.empty:
                    pw_msg.config(text="❌  Current password is wrong.", fg="#DC2626")
                    return
                df.loc[df["Username"]==uname, "Password"] = nw
                df.to_excel(USERS_FILE(), index=False, engine="openpyxl")
                cur_pw.delete(0,END)
                new_pw.delete(0,END)
                con_pw.delete(0,END)
                pw_msg.config(text="✅  Password changed successfully!", fg="#16A34A")
            except Exception as ex:
                pw_msg.config(text=f"❌  Error: {ex}", fg="#DC2626")

        make_button(pw_in, "🔑  Change Password", ACCENT_BLUE,
                    command=_change_password,
                    padx=14, pady=7).pack(anchor=W, pady=(6,0))

        # ── OPTION 2: About / Software Info ──────────────────────
        ab_card = Frame(body, bg="#F0FDF4",
                        highlightthickness=1, highlightbackground="#BBF7D0")
        ab_card.pack(fill=X, pady=(0,14))
        ab_in = Frame(ab_card, bg="#F0FDF4", padx=16, pady=12)
        ab_in.pack(fill=X)

        Label(ab_in, text="ℹ️  About This Software",
              font=("Segoe UI",10,"bold"),
              bg="#F0FDF4", fg="#15803D").pack(anchor=W)
        Label(ab_in,
              text=(
                  f"GST Visitor Management System\n"
                  f"Version      :  {CURRENT_VERSION}\n"
                  f"Developer   :  Udbhav K\n"
                  f"Department :  GST Department Karnataka\n"
                  f"Build          :  April 2026"
              ),
              font=("Segoe UI",9),
              bg="#F0FDF4", fg="#16A34A",
              justify=LEFT).pack(anchor=W, pady=(6,0))

        # ── OPTION 5: Check for Update ───────────────────────────
        upd_card = Frame(body, bg="#EFF6FF",
                         highlightthickness=1, highlightbackground="#BFDBFE")
        upd_card.pack(fill=X)
        upd_in = Frame(upd_card, bg="#EFF6FF", padx=16, pady=12)
        upd_in.pack(fill=X)

        Label(upd_in, text="⬆  Software Update",
              font=("Segoe UI",10,"bold"),
              bg="#EFF6FF", fg="#1E40AF").pack(anchor=W)
        Label(upd_in,
              text="Check if a newer version is available online.\n"
                   "Update downloads and installs automatically.",
              font=("Segoe UI",9),
              bg="#EFF6FF", fg="#1D4ED8",
              justify=LEFT).pack(anchor=W, pady=(4,8))

        make_button(upd_in, "⬆  Check for Update Now", ACCENT_BLUE,
                    command=lambda: [top.destroy(),
                                     check_for_update_async(root, manual=True)],
                    padx=14, pady=8).pack(anchor=W)

        # Close button
        Frame(top, bg="#E2E8F0", height=1).pack(fill=X)
        close_f = Frame(top, bg="#FFFFFF", pady=10); close_f.pack(fill=X)
        make_button(close_f, "✕  Close", BTN_CLEAR,
                    command=top.destroy,
                    padx=16, pady=8).pack(side=RIGHT, padx=16)

    def update_stats():
        def _do_read():
            try:
                df    = get_cached_df()   # uses cache — no Excel read if fresh
                today = datetime.now().strftime(DATE_FORMAT)
                df["Date"] = df["Date"].astype(str).str.strip()
                df["Out"]  = df["Out"].astype(str).str.strip()
                td    = df[df["Date"] == today]
                still = td[td["Out"].isin(OUT_EMPTY)]
                inside_count = 0
                for _, row in still.iterrows():
                    try:
                        ids = str(row.get("ID Cards","")).strip()
                        if ids and ids not in ("","nan","NaN","None"):
                            id_list = [x.strip() for x in
                                      ids.replace(";",",").split(",")
                                      if x.strip()]
                            inside_count += max(len(id_list), 1)
                        else:
                            val = str(row.get("Total Members","1")).strip()
                            if val in ("","nan","NaN","None","0"):
                                val = "1"
                            inside_count += int(float(val))
                    except:
                        inside_count += 1
                root.after(0, lambda t=len(td), i=inside_count: (
                    cv_s.set(f"📋  Today: {t}"),
                    iv_s.set(f"●  Inside: {i}")
                ))
            except Exception as _se:
                print(f"update_stats error: {_se}")
        threading.Thread(target=_do_read, daemon=True).start()
        root.after(5000, update_stats)   # every 5 seconds
    update_stats()

    # ── TODAY detail popup ──
    def show_today_detail():
        """Full visitor report popup — date filter, all data, save to PDF/Excel."""
        try:
            top = Toplevel(root)
            top.title("Visitor Report")
            top.geometry("1100x620")
            top.configure(bg=BG_PAGE)
            top.grab_set()
            apply_modern_style()

            today = datetime.now().strftime(DATE_FORMAT)

            # ── Header ──
            h = Frame(top, bg=BG_HEADER, pady=10); h.pack(fill=X)
            Label(h, text="  📋  Visitor Report",
                  font=("Segoe UI",12,"bold"),
                  bg=BG_HEADER, fg=TEXT_WHITE).pack(side=LEFT, pady=4)
            Button(h, text="✕ Close", font=("Segoe UI",9),
                   bg=BG_HEADER, fg="#93C5FD", relief=FLAT,
                   cursor="hand2", command=top.destroy).pack(side=RIGHT, padx=14)

            # ── Filter bar ──
            fbar = Frame(top, bg=BG_CARD,
                         highlightthickness=1, highlightbackground=BORDER_CLR)
            fbar.pack(fill=X, padx=12, pady=(8,0))
            fi = Frame(fbar, bg=BG_CARD, padx=12, pady=8); fi.pack(fill=X)

            Label(fi, text="From:", font=("Segoe UI",9,"bold"),
                  bg=BG_CARD, fg=TEXT_DARK).pack(side=LEFT)
            from_var = StringVar(value=today)
            Entry(fi, textvariable=from_var, width=12,
                  font=("Segoe UI",9), relief=FLAT,
                  bg=FIELD_BG, fg=TEXT_DARK,
                  highlightthickness=1,
                  highlightbackground=BORDER_CLR).pack(side=LEFT, padx=(4,12), ipady=4)

            Label(fi, text="To:", font=("Segoe UI",9,"bold"),
                  bg=BG_CARD, fg=TEXT_DARK).pack(side=LEFT)
            to_var = StringVar(value=today)
            Entry(fi, textvariable=to_var, width=12,
                  font=("Segoe UI",9), relief=FLAT,
                  bg=FIELD_BG, fg=TEXT_DARK,
                  highlightthickness=1,
                  highlightbackground=BORDER_CLR).pack(side=LEFT, padx=(4,12), ipady=4)

            Label(fi, text="dd-mm-yyyy", font=("Segoe UI",7,"italic"),
                  bg=BG_CARD, fg=TEXT_LIGHT).pack(side=LEFT, padx=(0,16))

            # Type filter
            Label(fi, text="Type:", font=("Segoe UI",9,"bold"),
                  bg=BG_CARD, fg=TEXT_DARK).pack(side=LEFT)
            type_var = StringVar(value="All")
            type_cb = ttk.Combobox(fi, textvariable=type_var, width=10,
                                   state="readonly",
                                   values=["All","Inside","Exited"])
            type_cb.pack(side=LEFT, padx=(4,12))

            # Result count label
            count_var = StringVar(value="")
            Label(fi, textvariable=count_var,
                  font=("Segoe UI",9,"bold"),
                  bg=BG_CARD, fg=ACCENT_BLUE).pack(side=LEFT, padx=(8,0))

            # ── Treeview ──
            cols = ["Date","Group ID","Arrival","Out","Visitor",
                    "Phone","ID Cards","Company","Purpose","Officer"]
            tree_frame = Frame(top, bg=BG_PAGE)
            tree_frame.pack(fill=BOTH, expand=True, padx=12, pady=6)
            tv = build_treeview(tree_frame, cols, height=16)

            _filtered_df = [None]   # store for PDF/Excel save

            def _apply_filter(*_):
                for row in tv.get_children(): tv.delete(row)
                try:
                    df_all = pd.read_excel(VISITORS_FILE(), dtype=str,
                                           engine="openpyxl").fillna("")
                    df_all["Date"] = df_all["Date"].astype(str).str.strip()
                    df_all["Out"]  = df_all["Out"].astype(str).str.strip()

                    # Date filter
                    try:
                        fd = datetime.strptime(from_var.get().strip(), DATE_FORMAT)
                        td2 = datetime.strptime(to_var.get().strip(), DATE_FORMAT)
                    except:
                        messagebox.showwarning("Date Format",
                            "Use dd-mm-yyyy format.\nExample: 15-04-2026",
                            parent=top)
                        return

                    mask = df_all["Date"].apply(
                        lambda d: _in_range(d, fd, td2))
                    df_f = df_all[mask].copy()

                    # Type filter
                    t_sel = type_var.get()
                    if t_sel == "Inside":
                        df_f = df_f[df_f["Out"].isin(OUT_EMPTY)]
                    elif t_sel == "Exited":
                        df_f = df_f[~df_f["Out"].isin(OUT_EMPTY)]

                    _filtered_df[0] = df_f
                    count_var.set(f"{len(df_f)} record(s)")

                    for _, r in df_f.iterrows():
                        out_v = str(r.get("Out","")).strip()
                        status = out_v if out_v not in OUT_EMPTY else "Inside ✅"
                        tv.insert("", END, values=[
                            str(r.get("Date","")),
                            str(r.get("Group ID","")),
                            str(r.get("Arrival","")),
                            status,
                            str(r.get("Visitor","")),
                            str(r.get("Phone","")),
                            str(r.get("ID Cards","")),
                            str(r.get("Company","")),
                            str(r.get("Purpose","")),
                            str(r.get("Officer",""))
                        ])
                except Exception as ex:
                    messagebox.showerror("Error", str(ex), parent=top)

            def _in_range(d_str, fd, td2):
                try:
                    d = datetime.strptime(d_str.strip(), DATE_FORMAT)
                    return fd <= d <= td2
                except: return False

            # Filter buttons
            make_button(fi, "🔍  Search", BTN_SEARCH,
                        command=_apply_filter,
                        padx=12, pady=6).pack(side=LEFT, padx=(8,4))

            def _save_excel():
                if _filtered_df[0] is None or _filtered_df[0].empty:
                    messagebox.showwarning("Empty",
                        "No data to save. Run Search first.", parent=top)
                    return
                from tkinter import filedialog
                fp = filedialog.asksaveasfilename(
                    parent=top,
                    title="Save Report as Excel",
                    defaultextension=".xlsx",
                    initialfile=f"GST_Report_{from_var.get()}_{to_var.get()}",
                    filetypes=[("Excel","*.xlsx"),("All","*.*")])
                if not fp: return
                try:
                    df_save = _filtered_df[0]
                    df_save.to_excel(fp, index=False, engine="openpyxl")
                    # Style it
                    from openpyxl import load_workbook
                    from openpyxl.styles import PatternFill,Font,Alignment,Border,Side
                    wb = load_workbook(fp); ws = wb.active
                    hf  = PatternFill("solid", fgColor="1E3A5F")
                    hft = Font(bold=True, color="FFFFFF",
                               name="Segoe UI", size=10)
                    th  = Side(style="thin", color="D1D5DB")
                    bd  = Border(left=th,right=th,top=th,bottom=th)
                    for c in ws[1]:
                        c.fill=hf; c.font=hft
                        c.alignment=Alignment(horizontal="center")
                        c.border=bd
                    for row in ws.iter_rows(min_row=2):
                        for c in row:
                            c.alignment=Alignment(horizontal="center")
                            c.border=bd
                            if c.row%2==0:
                                c.fill=PatternFill("solid",fgColor="EFF6FF")
                    for col in ws.columns:
                        ml=max((len(str(c.value)) if c.value else 0) for c in col)
                        ws.column_dimensions[col[0].column_letter].width=min(ml+4,30)
                    wb.save(fp)
                    if messagebox.askyesno("✅ Saved",
                        f"Saved {len(df_save)} records.\n\n{fp}\n\nOpen now?",
                        parent=top):
                        os.startfile(fp)
                except Exception as ex:
                    messagebox.showerror("Error", str(ex), parent=top)

            def _save_pdf():
                if _filtered_df[0] is None or _filtered_df[0].empty:
                    messagebox.showwarning("Empty",
                        "No data to save. Run Search first.", parent=top)
                    return
                from tkinter import filedialog
                fp = filedialog.asksaveasfilename(
                    parent=top,
                    title="Save Report as PDF",
                    defaultextension=".pdf",
                    initialfile=f"GST_Report_{from_var.get()}_{to_var.get()}",
                    filetypes=[("PDF","*.pdf"),("All","*.*")])
                if not fp: return
                try:
                    # Save as Excel first, then convert using openpyxl
                    # Pure Python PDF without reportlab
                    import tkinter as _tk
                    df_pdf = _filtered_df[0]
                    # Build HTML then save — simpler and works everywhere
                    html_fp = fp.replace(".pdf",".html")
                    html = ["<html><head><style>",
                            "body{font-family:Arial,sans-serif;font-size:11px}",
                            "h2{color:#1E3A5F}",
                            "table{border-collapse:collapse;width:100%}",
                            "th{background:#1E3A5F;color:white;padding:6px;text-align:center}",
                            "td{border:1px solid #D1D5DB;padding:5px;text-align:center}",
                            "tr:nth-child(even){background:#EFF6FF}",
                            "</style></head><body>",
                            f"<h2>GST Department — Visitor Report</h2>",
                            f"<p>Period: {from_var.get()} to {to_var.get()} | "
                            f"Total: {len(df_pdf)} records | "
                            f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}</p>",
                            "<table><tr>"]
                    show_cols = ["Date","Group ID","Arrival","Out","Visitor",
                                 "Phone","Company","Purpose","Officer"]
                    for c in show_cols:
                        html.append(f"<th>{c}</th>")
                    html.append("</tr>")
                    for _, r in df_pdf.iterrows():
                        html.append("<tr>")
                        for c in show_cols:
                            val = str(r.get(c,"")).strip()
                            if c == "Out" and val in OUT_EMPTY:
                                val = "Inside"
                            html.append(f"<td>{val}</td>")
                        html.append("</tr>")
                    html.append("</table></body></html>")
                    # Save HTML file
                    with open(html_fp,"w",encoding="utf-8") as hf_out:
                        hf_out.write("".join(html))
                    # Open in browser — user can Print → Save as PDF
                    import webbrowser
                    webbrowser.open(html_fp)
                    messagebox.showinfo("📄 Report Ready",
                        "Report opened in your browser.\n\n"
                        "To save as PDF:\n"
                        "  Press Ctrl+P → Change destination\n"
                        "  → 'Save as PDF' → Save\n\n"
                        f"HTML saved at: {html_fp}",
                        parent=top)
                except Exception as ex:
                    messagebox.showerror("Error", str(ex), parent=top)

            # Save buttons
            make_button(fi, "💾  Save Excel", BTN_IN,
                        command=_save_excel,
                        padx=12, pady=6).pack(side=LEFT, padx=4)
            make_button(fi, "📄  Save PDF", "#7C3AED",
                        command=_save_pdf,
                        padx=12, pady=6).pack(side=LEFT, padx=4)

            # Load today's data immediately
            _apply_filter()

        except Exception as ex:
            messagebox.showerror("Error", str(ex))

    # ── INSIDE detail popup ──
    def show_inside_detail():
        try:
            df    = get_cached_df()   # use cache — instant
            today = datetime.now().strftime(DATE_FORMAT)
            df["Date"] = df["Date"].astype(str).str.strip()
            df["Out"]  = df["Out"].astype(str).str.strip()
            td    = df[df["Date"] == today]
            still = td[td["Out"].isin(OUT_EMPTY)]
            top = Toplevel(root)
            top.title("Still Inside Now")
            top.geometry("820x400")
            top.configure(bg=BG_PAGE)
            top.grab_set()
            apply_modern_style()
            h = Frame(top, bg="#16A34A", pady=10); h.pack(fill=X)
            Label(h, text=f"  ●  Visitors Still Inside — {len(still)} group(s)",
                  font=("Segoe UI",11,"bold"), bg="#16A34A",
                  fg=TEXT_WHITE, pady=6).pack(side=LEFT)
            Button(h, text="✕ Close", font=("Segoe UI",9),
                   bg="#16A34A", fg=TEXT_WHITE, relief=FLAT,
                   cursor="hand2", command=top.destroy).pack(side=RIGHT, padx=14)
            if still.empty:
                Label(top, text="✅  No visitors inside right now.",
                      font=("Segoe UI",11,"bold"), bg=BG_PAGE,
                      fg="#16A34A").pack(pady=30)
                return
            cols = ["ID Cards","Visitor Name","Phone","Arrival","Company","Purpose","Remaining"]
            t = build_treeview(top, cols, height=10)
            for _, r in still.iterrows():
                t.insert("", END, values=[
                    str(r.get("ID Cards","")),
                    str(r.get("Visitor","")),
                    str(r.get("Phone","")),
                    str(r.get("Arrival","")),
                    str(r.get("Company","")),
                    str(r.get("Purpose","")),
                    str(r.get("Remaining",""))
                ])
            Label(top, text="💡  To exit: copy ID from above → paste in ID field → VISITOR OUT",
                  font=("Segoe UI",8,"italic"), bg=BG_PAGE,
                  fg=TEXT_MID).pack(pady=6)
        except Exception as ex:
            messagebox.showerror("Error", str(ex))

    # ── Build pills now (after callbacks defined) ──
    _spill(pills_f, cv_s,   "#EFF6FF", ACCENT_BLUE,  "📋", show_today_detail)
    _spill(pills_f, iv_s,   "#F0FDF4", ACCENT_GREEN, "●",  show_inside_detail)
    _spill(pills_f, cam_s,  "#F5F3FF", PURPLE,       "")
    _spill(pills_f, scan_s, "#FFF7ED", ACCENT_AMBER, "")

    def check_cam():
        idx = detect_webcam()
        cam_s.set("📷  Webcam: Ready ✅" if idx >= 0 else "📷  No webcam")
    threading.Thread(target=check_cam, daemon=True).start()

    def check_scanner():
        found, _ = detect_scanner()
        scan_s.set("🔌  Scanner: Ready ✅" if found else "🔌  Not connected")
        root.after(10000, lambda: threading.Thread(
            target=check_scanner, daemon=True).start())
    threading.Thread(target=check_scanner, daemon=True).start()

    # ── REFRESH button in stats bar ──
    def _manual_refresh():
        update_stats()
        status_var.set("🔄  Refreshed — Inside count updated")
    ref_btn = Button(sb_inner, text="🔄 Refresh",
                     font=("Segoe UI",8,"bold"),
                     bg="#E2E8F0", fg="#1E293B",
                     relief=FLAT, cursor="hand2",
                     padx=10, pady=5, bd=0,
                     activebackground="#CBD5E1",
                     command=_manual_refresh)
    ref_btn.pack(side=RIGHT, padx=(0,6))
    form_area = Frame(main, bg=FRM_BG)
    form_area.pack(fill=BOTH, expand=True, padx=16, pady=12)

    # ── State variables — declared early so all functions can use them ──
    status_var      = StringVar(value="Enter visitor details and ID card numbers")
    _is_returning   = [False]
    _existing_photo = [""]
    _banner_shown   = [False]

    # ── Returning visitor banner ──
    # Must be created BEFORE form_card but packed ONLY when needed
    ret_banner = Frame(form_area, bg="#FFFFFF",
                       highlightthickness=2,
                       highlightbackground=ACCENT_BLUE)

    def hide_banner():
        if _banner_shown[0]:
            ret_banner.pack_forget()
            _banner_shown[0] = False

    # White form card
    form_card = Frame(form_area, bg=CARD_W,
                      highlightthickness=1, highlightbackground=BORDER_CLR)
    form_card.pack(fill=BOTH, expand=True)

    # Scrollable form
    scroll_outer = Frame(form_card, bg=CARD_W)
    scroll_outer.pack(fill=BOTH, expand=True)
    form_canvas = Canvas(scroll_outer, bg=CARD_W, highlightthickness=0)
    form_scroll = ttk.Scrollbar(scroll_outer, orient=VERTICAL,
                                command=form_canvas.yview,
                                style="Modern.Vertical.TScrollbar")
    form_canvas.configure(yscrollcommand=form_scroll.set)
    form_scroll.pack(side=RIGHT, fill=Y)
    form_canvas.pack(side=LEFT, fill=BOTH, expand=True)
    fb = Frame(form_canvas, bg=CARD_W, padx=20, pady=16)
    fb_win = form_canvas.create_window((0,0), window=fb, anchor=NW)
    def _fcfg(e): form_canvas.itemconfig(fb_win, width=e.width)
    form_canvas.bind("<Configure>", _fcfg)
    def _fbcfg(e): form_canvas.configure(scrollregion=form_canvas.bbox("all"))
    fb.bind("<Configure>", _fbcfg)
    def _mw(e): form_canvas.yview_scroll(int(-1*(e.delta/120)),"units")
    form_canvas.bind_all("<MouseWheel>", _mw)

    entries = {}
    PAD = dict(padx=(0,16), pady=(0,10))

    def _sec(text):
        """Section divider — matches image style"""
        f = Frame(fb, bg=CARD_W)
        f.grid(row=_sec.row, column=0, columnspan=4, sticky=EW,
               padx=0, pady=(14,6))
        Label(f, text=text, font=("Segoe UI", 9, "bold"),
              bg=CARD_W, fg=ACCENT_BLUE).pack(side=LEFT)
        Frame(f, bg=BORDER_CLR, height=1).pack(side=LEFT, fill=X,
                                                expand=True, padx=(10,0))
        _sec.row += 1
    _sec.row = 0

    def _lbl(row, col, text, span=1):
        Label(fb, text=text, font=("Segoe UI", 9, "bold"),
              bg=CARD_W, fg=TEXT_MID, anchor=W).grid(
            row=row, column=col, columnspan=span,
            sticky=W, padx=(0,16), pady=(8,2))

    def _ent(row, col, key, span=1):
        e = make_entry(fb)
        e.grid(row=row, column=col, columnspan=span,
               sticky=EW, padx=(0,16), pady=(0,6), ipady=7)
        entries[key] = e
        return e

    for c in range(4): fb.columnconfigure(c, weight=1)

    # ── SECTION 1: Visitor Details ──
    _sec.row = 0
    _sec("Visitor Details")

    _lbl(_sec.row, 0, "📞  Phone Number")
    _lbl(_sec.row, 1, "👤  Visitor Name")
    _lbl(_sec.row, 2, "🏢  Company / Org")
    _lbl(_sec.row, 3, "📋  Purpose of Visit")
    _sec.row += 1

    _ent(_sec.row, 0, "Phone")
    _ent(_sec.row, 1, "Visitor Name")
    _ent(_sec.row, 2, "Company")
    purpose_var = StringVar()
    purpose_cb = ttk.Combobox(fb, textvariable=purpose_var,
                               values=load_purposes(),
                               font=("Segoe UI",10),
                               style="Modern.TCombobox")
    purpose_cb.grid(row=_sec.row, column=3, sticky=EW,
                    padx=(0,16), pady=(0,6), ipady=5)
    purpose_cb.set("Select purpose")
    entries["Purpose"] = purpose_cb
    _sec.row += 1

    # ── VIP Code row — shown only if VIP system is ON ──
    vip_row_f = Frame(fb, bg="#FFFBEB",
                      highlightthickness=1, highlightbackground="#C8A84B")
    vip_row_in = Frame(vip_row_f, bg="#FFFBEB", padx=10, pady=6)
    vip_row_in.pack(fill=X)
    Label(vip_row_in, text="⭐  VIP Code",
          font=("Segoe UI",8,"bold"),
          bg="#FFFBEB", fg="#92400E").pack(side=LEFT, padx=(0,8))
    vip_entry_var = StringVar()
    vip_entry_field = Entry(vip_row_in, textvariable=vip_entry_var,
                            font=("Segoe UI",10,"bold"),
                            bg="#FEF3C7", fg="#1B2E4B",
                            relief=FLAT, width=12,
                            highlightthickness=1,
                            highlightbackground="#C8A84B",
                            highlightcolor="#C8A84B",
                            insertbackground="#92400E")
    vip_entry_field.pack(side=LEFT, ipady=5, padx=(0,8))
    Label(vip_row_in, text="← Type VIP code + Tab to load all details (e.g. B7766)",
          font=("Segoe UI",8,"italic"),
          bg="#FFFBEB", fg="#92400E").pack(side=LEFT)
    vip_status_lbl = Label(vip_row_in, text="",
                           font=("Segoe UI",8,"bold"),
                           bg="#FFFBEB", fg="#16A34A")
    vip_status_lbl.pack(side=RIGHT, padx=8)

    # Show or hide VIP row based on admin setting
    if load_vip_enabled():
        vip_row_f.grid(row=_sec.row, column=0, columnspan=4,
                       sticky=EW, padx=(0,16), pady=(0,8))
        _sec.row += 1
    # If VIP is OFF — vip_row_f exists but not shown (so no NameError later)

    # ── SECTION 2: GST & Office Details ──
    _sec("GST & Office Details")

    _lbl(_sec.row, 0, "🔖  GST No")
    _lbl(_sec.row, 1, "🧑\u200d💼  Officer to Visit")
    _lbl(_sec.row, 2, "🏗  Block")
    _lbl(_sec.row, 3, "🔢  Floor No")
    _sec.row += 1

    # GST entry with validation
    gst_frame = Frame(fb, bg=CARD_W)
    gst_frame.grid(row=_sec.row, column=0, sticky=EW, padx=(0,16), pady=(0,2))
    gst_frame.columnconfigure(0, weight=1)
    gst_entry = make_entry(gst_frame)
    gst_entry.grid(row=0, column=0, sticky=EW, ipady=7)
    gst_hint_lbl = Label(gst_frame,
                         text="15 characters (e.g. 29ABCDE1234F1Z5)",
                         font=("Segoe UI",7,"italic"), bg=CARD_W, fg=TEXT_LIGHT)
    gst_hint_lbl.grid(row=1, column=0, sticky=W, pady=(2,4))

    def _gst_validate(*_):
        import re
        raw = gst_entry.get().upper()
        pos = gst_entry.index(INSERT)
        gst_entry.delete(0,END); gst_entry.insert(0,raw)
        try: gst_entry.icursor(min(pos,len(raw)))
        except: pass
        pat = r"^[0-3][0-9][A-Z]{5}[0-9]{4}[A-Z][1-9A-Z]Z[0-9A-Z]$"
        if   raw=="":           gst_entry.config(highlightbackground=BORDER_CLR); gst_hint_lbl.config(text="15 characters (e.g. 29ABCDE1234F1Z5)",fg=TEXT_LIGHT)
        elif re.match(pat,raw): gst_entry.config(highlightbackground=ACCENT_GREEN); gst_hint_lbl.config(text="✅  Valid GSTIN",fg=ACCENT_GREEN)
        elif len(raw)==15:      gst_entry.config(highlightbackground=ACCENT_RED); gst_hint_lbl.config(text="❌  Invalid",fg=ACCENT_RED)
        else:                   gst_entry.config(highlightbackground=ACCENT_AMBER); gst_hint_lbl.config(text=f"⏳  {len(raw)}/15",fg=ACCENT_AMBER)
    gst_entry.bind("<KeyRelease>",_gst_validate)
    gst_entry.bind("<FocusOut>",  _gst_validate)
    entries["GST No"] = gst_entry

    # Officer autocomplete
    off_frame = Frame(fb, bg=CARD_W)
    off_frame.grid(row=_sec.row, column=1, sticky=EW, padx=(0,16), pady=(0,2))
    off_frame.columnconfigure(0, weight=1)
    officer_var = StringVar()
    officer_entry = make_entry(off_frame, textvariable=officer_var)
    officer_entry.grid(row=0, column=0, sticky=EW, ipady=7)
    off_hint = Label(off_frame, text="Type code or name — fields auto-fill",
                     font=("Segoe UI",7,"italic"), bg=CARD_W, fg=TEXT_LIGHT)
    off_hint.grid(row=1, column=0, sticky=W, pady=(2,4))
    entries["Officer"] = officer_entry

    off_popup = None
    def _officer_autofill(query):
        rec = find_officer(query)
        if not rec:
            off_hint.config(text="Officer not found — fill manually", fg=ACCENT_AMBER); return
        entries["Division"].delete(0,END); entries["Division"].insert(0,rec.get("Division",""))
        entries["Room No"].delete(0,END);  entries["Room No"].insert(0,rec.get("Room No",""))
        bv=rec.get("Block","").strip()
        if bv: entries["Block"].set(bv)
        fv=rec.get("Floor","").strip()
        if fv: entries["Floor"].set(fv)
        off_hint.config(text=f"✅  {rec.get('Officer Name','')}  •  {rec.get('Division','')}",fg=ACCENT_GREEN)

    def _show_off_dd(matches):
        nonlocal off_popup
        _hide_off_dd()
        if not matches: return
        off_popup = Toplevel(root)
        off_popup.overrideredirect(True)
        off_popup.configure(bg=CARD_W)
        x=officer_entry.winfo_rootx(); y=officer_entry.winfo_rooty()+officer_entry.winfo_height()+2
        w=max(officer_entry.winfo_width(),320)
        off_popup.geometry(f"{w}x{min(len(matches)*36,180)}+{x}+{y}")
        off_popup.attributes("-topmost",True)
        lb=Listbox(off_popup,font=("Segoe UI",9),bg=CARD_W,fg=TEXT_DARK,
                   selectbackground=ACCENT_BLUE,selectforeground=TEXT_WHITE,
                   relief=FLAT,highlightthickness=0,activestyle="none",height=min(len(matches),5))
        lb.pack(fill=BOTH,expand=True)
        for m in matches:
            lb.insert(END,f"{m.get('Officer Code','')}  —  {m.get('Officer Name','')}  ({m.get('Designation','')})")
        def _sel(ev=None):
            sel=lb.curselection()
            if not sel: return
            chosen=matches[sel[0]]
            officer_var.set(chosen.get("Officer Name",""))
            _officer_autofill(chosen.get("Officer Code",""))
            _hide_off_dd(); officer_entry.focus_set()
        lb.bind("<ButtonRelease-1>",_sel); lb.bind("<Return>",_sel)

    def _hide_off_dd(*_):
        nonlocal off_popup
        if off_popup:
            try: off_popup.destroy()
            except: pass
            off_popup=None

    def _on_off_key(*_):
        q=officer_var.get().strip()
        if len(q)<1: _hide_off_dd(); return
        m=search_officers(q)
        if m: _show_off_dd(m)
        else: _hide_off_dd()

    def _on_off_focusout(ev):
        root.after(200,_hide_off_dd)
        q=officer_var.get().strip()
        if q: _officer_autofill(q)

    officer_var.trace_add("write",_on_off_key)
    officer_entry.bind("<FocusOut>",_on_off_focusout)
    officer_entry.bind("<Tab>",lambda ev:_officer_autofill(officer_var.get()))
    officer_entry.bind("<Escape>",_hide_off_dd)

    # Block + Floor dropdowns
    block_var = StringVar(value="")
    block_cb = ttk.Combobox(fb, textvariable=block_var,
                             values=["","A Block","B Block"],
                             state="readonly", font=("Segoe UI",10),
                             style="Modern.TCombobox")
    block_cb.grid(row=_sec.row, column=2, sticky=EW, padx=(0,16), pady=(0,6), ipady=5)
    entries["Block"] = block_cb

    floor_var = StringVar(value="")
    floor_cb = ttk.Combobox(fb, textvariable=floor_var,
                             values=["","Ground Floor","1st Floor","2nd Floor",
                                     "3rd Floor","4th Floor","5th Floor","6th Floor"],
                             state="readonly", font=("Segoe UI",10),
                             style="Modern.TCombobox")
    floor_cb.grid(row=_sec.row, column=3, sticky=EW, padx=(0,16), pady=(0,6), ipady=5)
    entries["Floor"] = floor_cb
    _sec.row += 1

    # ── Division + Room No ──
    _lbl(_sec.row, 0, "🏛  Division")
    _lbl(_sec.row, 1, "🚪  Room No")
    _sec.row += 1
    _ent(_sec.row, 0, "Division")
    _ent(_sec.row, 1, "Room No")
    _sec.row += 1

    # ── Remarks ──
    _lbl(_sec.row, 0, "📝  Remarks", span=4)
    _sec.row += 1
    rem_e = make_entry(fb)
    rem_e.grid(row=_sec.row, column=0, columnspan=4, sticky=EW,
               padx=(0,16), pady=(0,8), ipady=7)
    entries["Remarks"] = rem_e
    _sec.row += 1

    # ── ID CARD NUMBERS ──
    Frame(fb, bg=BORDER_CLR, height=1).grid(
        row=_sec.row, column=0, columnspan=4, sticky=EW, pady=(4,6))
    _sec.row += 1

    id_hdr = Frame(fb, bg="#FFFDE7",
                   highlightthickness=1, highlightbackground=ACCENT_AMBER)
    id_hdr.grid(row=_sec.row, column=0, columnspan=4, sticky=EW, pady=(0,6))
    id_hdr_in = Frame(id_hdr, bg="#FFFDE7", padx=12, pady=8)
    id_hdr_in.pack(fill=X)
    Label(id_hdr_in,
          text="🪪  ID CARD NUMBERS — Enter IDs separated by commas:  A101, A102, A103",
          font=("Segoe UI",8,"bold"), bg="#FFFDE7", fg=ACCENT_AMBER).pack(side=LEFT)
    _sec.row += 1

    id_var = StringVar()
    id_entry = Entry(fb, textvariable=id_var,
                     font=("Segoe UI",11), bg="#FFFDE7", fg=TEXT_DARK,
                     relief=FLAT, highlightthickness=2,
                     highlightbackground=ACCENT_AMBER, highlightcolor=ACCENT_BLUE,
                     insertbackground=ACCENT_BLUE)
    id_entry.grid(row=_sec.row, column=0, columnspan=4, sticky=EW,
                  padx=(0,16), pady=(0,6), ipady=10)
    _sec.row += 1

    pc_frame = Frame(fb, bg=CARD_W)
    pc_frame.grid(row=_sec.row, column=0, columnspan=4, sticky=W, pady=(0,4))
    Label(pc_frame, text="No. of Persons :",
          font=("Segoe UI",9,"bold"), bg=CARD_W, fg=TEXT_MID).pack(side=LEFT)
    persons_var = StringVar(value="0")
    Label(pc_frame, textvariable=persons_var,
          font=("Segoe UI",14,"bold"), bg=STATUS_BG,
          fg=BG_HEADER, padx=14, pady=2).pack(side=LEFT, padx=8)
    Label(pc_frame, text="(auto from ID count)",
          font=("Segoe UI",8,"italic"), bg=CARD_W, fg=TEXT_LIGHT).pack(side=LEFT)
    _sec.row += 1

    id_list_lbl = Label(fb, text="IDs will appear here",
                        font=("Segoe UI",8,"italic"), bg="#F0F4FF",
                        fg=TEXT_LIGHT, wraplength=800, justify=LEFT,
                        padx=10, pady=6)
    id_list_lbl.grid(row=_sec.row, column=0, columnspan=4,
                     sticky=EW, padx=(0,16), pady=(0,10))
    _sec.row += 1

    def parse_ids():
        return list(dict.fromkeys(
            x.strip() for x in id_var.get().replace(";",",").split(",") if x.strip()))

    def on_id_change(*args):
        ids = parse_ids()
        persons_var.set(str(len(ids)))
        if ids:
            id_list_lbl.config(
                text="  ✅  "+"   |   ".join(f"[{i+1}] {d}" for i,d in enumerate(ids)),
                fg=ACCENT_GREEN, bg="#F0FDF4")
        else:
            id_list_lbl.config(text="IDs will appear here", fg=TEXT_LIGHT, bg="#F0F4FF")
    id_var.trace_add("write", on_id_change)

    def show_returning_banner(name, phone, company, gst,
                               purpose, officer, division, room,
                               visits, photo_path, vip_code=""):
        """Show returning visitor banner above form."""
        # Clear previous content
        for w in ret_banner.winfo_children():
            w.destroy()

        # ── Header ──
        hdr = Frame(ret_banner, bg=ACCENT_BLUE, pady=8)
        hdr.pack(fill=X)
        hdr_l = Frame(hdr, bg=ACCENT_BLUE); hdr_l.pack(side=LEFT, padx=14)
        Label(hdr_l, text="● Returning Visitor",
              font=("Segoe UI",9,"bold"),
              bg=ACCENT_BLUE, fg="#FFFFFF").pack(side=LEFT)
        if vip_code:
            Label(hdr_l, text=f"  ⭐ VIP: {vip_code}",
                  font=("Segoe UI",9,"bold"),
                  bg=ACCENT_BLUE, fg="#FEF3C7").pack(side=LEFT)
        hdr_r = Frame(hdr, bg=ACCENT_BLUE); hdr_r.pack(side=RIGHT, padx=14)
        Label(hdr_r, text=f"Phone {phone} matched — {visits} visit(s)",
              font=("Segoe UI",8), bg=ACCENT_BLUE, fg="#BFDBFE").pack(side=LEFT)
        Button(hdr_r, text=" ✕ ", font=("Segoe UI",9,"bold"),
               bg=ACCENT_BLUE, fg="#BFDBFE", relief=FLAT,
               cursor="hand2", command=hide_banner, bd=0,
               activebackground="#1D4ED8").pack(side=LEFT, padx=(8,0))

        # ── Body ──
        body = Frame(ret_banner, bg="#FFFFFF")
        body.pack(fill=X)

        # Photo column
        ph_col = Frame(body, bg="#EFF6FF", width=88)
        ph_col.pack(side=LEFT, fill=Y); ph_col.pack_propagate(False)
        ph_in  = Frame(ph_col, bg="#EFF6FF"); ph_in.pack(expand=True, pady=10)

        loaded = False
        # photo_path can be a file path OR folder path — handle both
        photo_dir = ""
        if photo_path and photo_path not in ("","nan","NaN","None"):
            if os.path.isdir(photo_path):
                photo_dir = photo_path
            elif os.path.isfile(photo_path):
                photo_dir = os.path.dirname(photo_path)

        if photo_dir:
            try:
                imgs = [f for f in os.listdir(photo_dir)
                        if f.lower().endswith((".jpg",".jpeg",".png"))]
                if imgs:
                    from PIL import Image as _PI, ImageTk as _IT
                    img = _PI.open(os.path.join(photo_dir, imgs[0]))
                    img = img.resize((62,72), _PI.LANCZOS)
                    _tk_img = _IT.PhotoImage(img)
                    lbl = Label(ph_in, image=_tk_img, bg="#EFF6FF")
                    lbl.image = _tk_img
                    lbl.pack()
                    loaded = True
            except Exception as _ph_err:
                print(f"Photo load error: {_ph_err}")

        if not loaded:
            cv = Canvas(ph_in, width=62, height=72, bg="#DBEAFE",
                        highlightthickness=1, highlightbackground=ACCENT_BLUE)
            cv.pack()
            cv.create_text(31,36, text="👤",
                           font=("Segoe UI Emoji",28), fill=ACCENT_BLUE)

        Label(ph_in, text=f"🔁 {visits} visits",
              font=("Segoe UI",7,"bold"),
              bg="#EFF6FF", fg=ACCENT_BLUE).pack(pady=(4,0))

        # Info grid
        info = Frame(body, bg="#FFFFFF")
        info.pack(side=LEFT, fill=BOTH, expand=True, padx=12, pady=8)
        for c in range(4): info.columnconfigure(c, weight=1)

        def _cell(r, c, lbl, val, col="#1E293B"):
            Label(info, text=lbl, font=("Segoe UI",6,"bold"),
                  bg="#FFFFFF", fg="#94A3B8").grid(
                row=r*2, column=c, sticky=W, padx=(0,10))
            Label(info, text=val or "—", font=("Segoe UI",9,"bold"),
                  bg="#FFFFFF", fg=col).grid(
                row=r*2+1, column=c, sticky=W, padx=(0,10), pady=(0,4))

        _cell(0,0,"NAME",     name,    "#1E293B")
        _cell(0,1,"MOBILE",   phone,   ACCENT_BLUE)
        _cell(0,2,"COMPANY",  company, "#1E293B")
        _cell(0,3,"PURPOSE",  purpose, "#1E293B")
        _cell(1,0,"GST NO",   gst,     "#1E293B")
        _cell(1,1,"OFFICER",  officer, "#1E293B")
        _cell(1,2,"DIV/ROOM", f"{division} Rm{room}" if division else room, "#1E293B")
        _cell(1,3,"VIP CODE", f"⭐ {vip_code}" if vip_code else "—",
              "#92400E" if vip_code else "#94A3B8")

        # ── Strip ──
        strip = Frame(ret_banner, bg="#F0FDF4")
        strip.pack(fill=X)
        Frame(strip, bg="#BBF7D0", height=1).pack(fill=X)
        Label(strip, text="⚡  All fields auto-filled — check and click VISITOR IN",
              font=("Segoe UI",8,"bold"),
              bg="#F0FDF4", fg="#15803D",
              padx=14, pady=5).pack(side=LEFT)

        # ── Show banner — unpack form_card, show banner, repack form_card ──
        if not _banner_shown[0]:
            form_card.pack_forget()
            ret_banner.pack(fill=X, pady=(0,4))
            form_card.pack(fill=BOTH, expand=True)
            _banner_shown[0] = True

    def _show_vip_generate_popup(name, phone, company, visits, photo_path=""):
        """VIP code generation popup — matches reference image exactly."""
        vw = Toplevel(root)
        vw.title("Generate VIP Code")
        vw.geometry("480x620")
        vw.resizable(False, False)
        vw.configure(bg="#FFFFFF")
        # Do NOT use grab_set — it freezes main window scroll
        vw.transient(root)
        vw.focus_set()

        # ── Scrollable body ──
        outer_c = Canvas(vw, bg="#FFFFFF", highlightthickness=0)
        outer_c.pack(fill=BOTH, expand=True, side=LEFT)
        vsb = ttk.Scrollbar(vw, orient=VERTICAL, command=outer_c.yview,
                            style="Modern.Vertical.TScrollbar")
        vsb.pack(side=RIGHT, fill=Y)
        outer_c.configure(yscrollcommand=vsb.set)
        body = Frame(outer_c, bg="#FFFFFF")
        _bw = outer_c.create_window((0,0), window=body, anchor=NW)
        outer_c.bind("<Configure>", lambda e: outer_c.itemconfig(_bw, width=e.width))
        body.bind("<Configure>", lambda e: outer_c.configure(
            scrollregion=outer_c.bbox("all")))
        # Bind mousewheel ONLY to this popup canvas
        def _vip_scroll(e):
            outer_c.yview_scroll(int(-1*(e.delta/120)),"units")
        outer_c.bind("<MouseWheel>", _vip_scroll)
        body.bind("<MouseWheel>", _vip_scroll)
        # Restore main scroll on close
        def _on_vw_close():
            vw.destroy()
        vw.protocol("WM_DELETE_WINDOW", _on_vw_close)

        PAD = 24

        # ── Visitor info card ──
        vc = Frame(body, bg="#F8FAFF",
                   highlightthickness=1, highlightbackground="#E2E8F0")
        vc.pack(fill=X, padx=PAD, pady=(20,0))
        vc_in = Frame(vc, bg="#F8FAFF", padx=16, pady=14)
        vc_in.pack(fill=X)

        # Photo circle
        ph_cv = Canvas(vc_in, width=52, height=52, bg="#EFF6FF",
                       highlightthickness=1, highlightbackground="#BFDBFE")
        ph_cv.pack(side=LEFT, padx=(0,14))
        # Try load real photo — handle both file path and folder path
        photo_loaded = False
        _photo_dir = ""
        if photo_path and photo_path not in ("","nan","NaN","None"):
            if os.path.isdir(photo_path):
                _photo_dir = photo_path
            elif os.path.isfile(photo_path):
                _photo_dir = os.path.dirname(photo_path)
        if _photo_dir:
            try:
                imgs = [f for f in os.listdir(_photo_dir)
                        if f.lower().endswith((".jpg",".jpeg",".png"))]
                if imgs:
                    from PIL import Image as PILImage, ImageTk as PILTk
                    img = PILImage.open(os.path.join(_photo_dir, imgs[0]))
                    img = img.resize((52,52), PILImage.LANCZOS)
                    _ph_img = PILTk.PhotoImage(img)
                    ph_cv.create_image(26,26,image=_ph_img)
                    ph_cv._img = _ph_img
                    photo_loaded = True
            except Exception as _pe:
                print(f"VIP popup photo error: {_pe}")
        if not photo_loaded:
            ph_cv.create_text(26,26,text="👤",
                              font=("Segoe UI Emoji",26), fill="#2563EB")

        vt = Frame(vc_in, bg="#F8FAFF"); vt.pack(side=LEFT, fill=X, expand=True)
        Label(vt, text=name, font=("Segoe UI",13,"bold"),
              bg="#F8FAFF", fg="#1E293B").pack(anchor=W)
        Label(vt, text=f"{company}  |  {phone}",
              font=("Segoe UI",9), bg="#F8FAFF", fg="#64748B").pack(anchor=W, pady=(2,0))
        Label(vt, text=f"📋  {visits} visits — VIP code recommended",
              font=("Segoe UI",9,"bold"), bg="#F8FAFF", fg="#2563EB").pack(anchor=W, pady=(4,0))

        # ── Type selection label ──
        Label(body, text="Select visitor type to auto-generate code:",
              font=("Segoe UI",10,"bold"), bg="#FFFFFF",
              fg="#1E293B").pack(anchor=W, padx=PAD, pady=(18,10))

        TYPES = [
            ("C", "CA / Auditor",      "Chartered Accountant"),
            ("A", "Advocate",          "Lawyer / Legal"),
            ("B", "Business / Trader", "Business Owner"),
            ("T", "Tax Consultant",    "Tax / Finance"),
            ("G", "Govt Officer",      "Government Dept"),
            ("R", "Regular",           "Frequent Visitor"),
        ]
        selected_type = [None]
        type_frames   = []
        prefix_cvs    = []   # canvas for circle
        label_refs    = []   # (name_lbl, desc_lbl)
        code_var      = StringVar(value="— select type above —")

        def _recolor_all():
            for i,(tf2,pcv,(nl,dl)) in enumerate(
                    zip(type_frames, prefix_cvs, label_refs)):
                if TYPES[i][0] == selected_type[0]:
                    # Gold selected style
                    tf2.config(bg="#FFFBEB",
                               highlightbackground="#C8A84B",
                               highlightthickness=2)
                    pcv.config(bg="#C8A84B")
                    pcv.delete("all")
                    pcv.create_oval(2,2,38,38, fill="#C8A84B", outline="")
                    pcv.create_text(20,20, text=TYPES[i][0],
                                    font=("Segoe UI",13,"bold"), fill="#1B2E4B")
                    nl.config(bg="#FFFBEB", fg="#92400E",
                              text=TYPES[i][1]+" ✓")
                    dl.config(bg="#FFFBEB", fg="#B45309")
                    for w in tf2.winfo_children():
                        if isinstance(w, Frame):
                            w.config(bg="#FFFBEB")
                else:
                    tf2.config(bg="#FFFFFF",
                               highlightbackground="#E2E8F0",
                               highlightthickness=1)
                    pcv.config(bg="#1B2E4B")
                    pcv.delete("all")
                    pcv.create_oval(2,2,38,38, fill="#1B2E4B", outline="")
                    pcv.create_text(20,20, text=TYPES[i][0],
                                    font=("Segoe UI",13,"bold"), fill="#C8A84B")
                    nl.config(bg="#FFFFFF", fg="#1E293B",
                              text=TYPES[i][1])
                    dl.config(bg="#FFFFFF", fg="#94A3B8")
                    for w in tf2.winfo_children():
                        if isinstance(w, Frame):
                            w.config(bg="#FFFFFF")

        def _select_type(idx):
            selected_type[0] = TYPES[idx][0]
            last4 = phone[-4:] if len(phone)>=4 else phone
            code_var.set(f"{TYPES[idx][0]}  {last4[0]}  {last4[1]}  {last4[2]}  {last4[3]}"
                         if len(last4)==4 else f"{TYPES[idx][0]}{last4}")
            _recolor_all()

        # ── 2-column type grid ──
        tgrid = Frame(body, bg="#FFFFFF")
        tgrid.pack(fill=X, padx=PAD, pady=(0,4))
        tgrid.columnconfigure(0, weight=1)
        tgrid.columnconfigure(1, weight=1)

        for idx,(prefix,label,desc) in enumerate(TYPES):
            ri, ci = divmod(idx,2)
            tf2 = Frame(tgrid, bg="#FFFFFF",
                        highlightthickness=1, highlightbackground="#E2E8F0",
                        cursor="hand2")
            tf2.grid(row=ri, column=ci,
                     padx=(0,8) if ci==0 else (0,0),
                     pady=(0,10), sticky=EW)
            type_frames.append(tf2)

            tf2_in = Frame(tf2, bg="#FFFFFF", padx=14, pady=12)
            tf2_in.pack(fill=X)

            # Circle canvas
            pcv = Canvas(tf2_in, width=40, height=40, bg="#1B2E4B",
                         highlightthickness=0)
            pcv.pack(side=LEFT, padx=(0,12))
            pcv.create_oval(2,2,38,38, fill="#1B2E4B", outline="")
            pcv.create_text(20,20, text=prefix,
                            font=("Segoe UI",13,"bold"), fill="#C8A84B")
            prefix_cvs.append(pcv)

            tf2_txt = Frame(tf2_in, bg="#FFFFFF")
            tf2_txt.pack(side=LEFT, fill=X, expand=True)
            nl = Label(tf2_txt, text=label,
                       font=("Segoe UI",10,"bold"),
                       bg="#FFFFFF", fg="#1E293B")
            nl.pack(anchor=W)
            dl = Label(tf2_txt, text=desc,
                       font=("Segoe UI",8),
                       bg="#FFFFFF", fg="#94A3B8")
            dl.pack(anchor=W)
            label_refs.append((nl,dl))

            # Bind all children
            _idx = idx
            for w in [tf2,tf2_in,pcv,tf2_txt,nl,dl]:
                w.bind("<Button-1>",
                       lambda e, i=_idx: _select_type(i))

        # ── Generated code box ──
        code_box = Frame(body, bg="#FFFBEB",
                         highlightthickness=2,
                         highlightbackground="#C8A84B")
        code_box.pack(fill=X, padx=PAD, pady=(6,16))
        code_in = Frame(code_box, bg="#FFFBEB", padx=18, pady=14)
        code_in.pack(fill=X)

        # Left: text
        code_left = Frame(code_in, bg="#FFFBEB"); code_left.pack(side=LEFT, fill=X, expand=True)
        Label(code_left, text="Auto-generated VIP Code",
              font=("Segoe UI",9,"bold"),
              bg="#FFFBEB", fg="#B45309").pack(anchor=W)
        code_disp = Label(code_left, textvariable=code_var,
                          font=("Segoe UI",28,"bold"),
                          bg="#FFFBEB", fg="#1B2E4B")
        code_disp.pack(anchor=W, pady=(6,4))
        last4 = phone[-4:] if len(phone)>=4 else phone
        Label(code_left,
              text=f"Prefix = visitor type  |  {last4} = last 4 of mobile",
              font=("Segoe UI",8,"italic"),
              bg="#FFFBEB", fg="#92400E").pack(anchor=W)

        # Right: star circle
        star_cv = Canvas(code_in, width=48, height=48,
                         bg="#C8A84B", highlightthickness=0)
        star_cv.pack(side=RIGHT, padx=(12,0))
        star_cv.create_oval(0,0,48,48, fill="#C8A84B", outline="")
        star_cv.create_text(24,24, text="⭐",
                            font=("Segoe UI Emoji",22), fill="#FFFFFF")

        # ── Save + Skip buttons ──
        btn_f = Frame(body, bg="#FFFFFF")
        btn_f.pack(fill=X, padx=PAD, pady=(0,24))

        def save_vip_code():
            if not selected_type[0]:
                messagebox.showwarning("Select Type",
                    "Please select visitor type first.", parent=vw)
                return
            last4 = phone[-4:] if len(phone)>=4 else phone
            code  = f"{selected_type[0]}{last4}"
            try:
                df = pd.read_excel(VISITORS_FILE(), dtype=str)
                if "VIP_Code" not in df.columns:
                    df["VIP_Code"] = ""
                df["Phone"] = df["Phone"].astype(str).str.strip()
                df.loc[df["Phone"]==phone, "VIP_Code"] = code
                df.to_excel(VISITORS_FILE(), index=False, engine='openpyxl')
                invalidate_cache()
                messagebox.showinfo("✅  VIP Code Saved",
                    f"VIP Code  :  {code}\n"
                    f"Visitor   :  {name}\n"
                    f"Phone     :  {phone}\n\n"
                    f"Next visit: type phone number\n"
                    f"→ VIP profile loads instantly ✅",
                    parent=vw)
                vw.destroy()
            except Exception as ex:
                messagebox.showerror("Error", str(ex), parent=vw)

        Button(btn_f, text="✅  Save VIP Code",
               font=("Segoe UI",11,"bold"),
               bg="#16A34A", fg="#FFFFFF",
               relief=FLAT, cursor="hand2",
               activebackground="#15803D",
               pady=12, bd=0,
               command=save_vip_code).pack(side=LEFT, fill=X, expand=True, padx=(0,10))
        Button(btn_f, text="Skip for now",
               font=("Segoe UI",10),
               bg="#F1F5F9", fg="#64748B",
               relief=FLAT, cursor="hand2",
               activebackground="#E2E8F0",
               pady=12, bd=0,
               command=vw.destroy).pack(side=LEFT, ipadx=10)

    def _fill_from_vip(vip_code):
        """Fill all fields from VIP code — leave Officer+Purpose blank."""
        vip_code = vip_code.strip().upper()
        if not vip_code:
            return
        try:
            df = get_cached_df()   # cache
            if "VIP_Code" not in df.columns:
                vip_status_lbl.config(
                    text="❌ No VIP records found", fg="#DC2626"); return
            df["VIP_Code"] = df["VIP_Code"].astype(str).str.strip().str.upper()
            matches = df[df["VIP_Code"] == vip_code]
            if matches.empty:
                vip_status_lbl.config(
                    text=f"❌ VIP code {vip_code} not found", fg="#DC2626"); return

            last = matches.iloc[-1]
            def _g(col):
                v = str(last.get(col,"")).strip()
                return "" if v in ("nan","NaN","None","") else v

            # Fill phone
            phone = _g("Phone")
            if phone:
                entries["Phone"].delete(0,END)
                entries["Phone"].insert(0, phone)

            # Fill name
            vname = _g("Visitor")
            if vname:
                entries["Visitor Name"].delete(0,END)
                entries["Visitor Name"].insert(0, vname)

            # Fill company
            company = _g("Company")
            if company:
                entries["Company"].delete(0,END)
                entries["Company"].insert(0, company)

            # Fill GST
            gst = _g("GST No")
            if gst:
                entries["GST No"].delete(0,END)
                entries["GST No"].insert(0, gst)

            # Fill Block/Floor
            blk = _g("Block")
            if blk and blk in ("A Block","B Block"):
                entries["Block"].set(blk)
            flr = _g("Floor")
            if flr:
                entries["Floor"].set(flr)

            # Set returning visitor flags
            photo_p = _g("Photo")
            _is_returning[0] = True
            _existing_photo[0] = photo_p

            # Officer + Purpose = BLANK (visitor will say today)
            # Division + Room = auto from officer if typed later

            visits = len(matches)
            vip_status_lbl.config(
                text=f"✅ VIP loaded — {vname} — {visits} visits",
                fg="#16A34A")

            status_var.set(
                f"⭐ VIP {vip_code} loaded — {vname} — ask Officer + Purpose")

            # Show banner
            show_returning_banner(
                vname, phone, company, gst,
                "",  # purpose blank
                "",  # officer blank
                _g("Division"), _g("Room No"),
                visits, photo_p, vip_code)

        except Exception as ex:
            vip_status_lbl.config(
                text=f"Error: {ex}", fg="#DC2626")

    def _on_vip_tab(event):
        """Triggered when Tab pressed in VIP field."""
        code = vip_entry_var.get().strip()
        if code:
            _fill_from_vip(code)

    vip_entry_field.bind("<FocusOut>", _on_vip_tab)
    vip_entry_field.bind("<Return>",   _on_vip_tab)

    def autofill_phone(event):
        phone = entries["Phone"].get().strip()
        if not phone: return
        _is_returning[0] = False
        _existing_photo[0] = ""
        hide_banner()
        try:
            df = get_cached_df()   # use cache — instant, no Excel read
            old = df[df["Phone"].astype(str).str.strip()==phone]
            if not old.empty:
                last    = old.iloc[-1]
                visits  = len(old)
                vname   = str(last.get("Visitor","")).strip()
                company = str(last.get("Company","")).strip()
                gst     = str(last.get("GST No","")).strip()
                purpose = str(last.get("Purpose","")).strip()
                officer = str(last.get("Officer","")).strip()
                division= str(last.get("Division","")).strip()
                room    = str(last.get("Room No","")).strip()

                # ── Photo: search ALL visits to find first valid photo ──
                # (last visit may not have photo, but 1st visit always does)
                photo_p = ""
                for _, visit_row in old.iterrows():
                    raw_p = str(visit_row.get("Photo","")).strip()
                    if raw_p in ("","nan","NaN","None"): continue
                    # Resolve: if file path → get folder, if folder → use directly
                    if os.path.isdir(raw_p):
                        photo_p = raw_p; break
                    elif os.path.isfile(raw_p):
                        photo_p = os.path.dirname(raw_p); break
                    elif os.path.isdir(os.path.dirname(raw_p)):
                        # path stored but file deleted — use folder
                        photo_p = os.path.dirname(raw_p); break

                # Safe VIP code read — handles missing column and nan values
                if "VIP_Code" in df.columns:
                    _vc = str(last.get("VIP_Code","")).strip()
                    vip_code = "" if _vc in ("nan","NaN","None","") else _vc
                else:
                    vip_code = ""


                # Auto-fill SAFE fields only: Name, Company, GST No
                # Purpose, Officer, Division, Room, Block, Floor — ALL BLANK
                # Receptionist must ask visitor each visit
                if vname and not entries["Visitor Name"].get().strip():
                    entries["Visitor Name"].delete(0, END)
                    entries["Visitor Name"].insert(0, vname)

                for col, key in [("Company","Company"), ("GST No","GST No")]:
                    if not entries[key].get().strip():
                        v = str(last.get(col,"")).strip()
                        if v and v not in ("","nan","NaN","None"):
                            entries[key].delete(0, END)
                            entries[key].insert(0, v)

                # Purpose LEFT BLANK — receptionist asks visitor each time
                # Officer, Division, Room, Block, Floor — LEFT BLANK always

                _is_returning[0] = True
                _existing_photo[0] = photo_p if photo_p not in ("","nan","NaN","None") else ""

                # VIP badge label
                vip_label = f"  ⭐ VIP: {vip_code}" if vip_code and vip_code not in ("","nan","NaN","None") else ""
                status_var.set(
                    f"✅  Returning visitor — {visits} visit(s){vip_label} — details auto-filled")

                # Show landscape banner
                show_returning_banner(
                    vname, phone, company, gst, purpose,
                    officer, division, room, visits,
                    _existing_photo[0], vip_code)

                # Show VIP prompt ONLY on 3rd+ visit AND no code assigned yet
                # Once code saved — NEVER shows again (permanent)
                has_valid_vip = bool(vip_code)
                if visits >= 3 and not has_valid_vip and load_vip_enabled():
                    if messagebox.askyesno(
                        "⭐ Generate VIP Code?",
                        f"Visitor  :  {vname}\n"
                        f"Phone    :  {phone}\n"
                        f"Visits   :  {visits} times\n\n"
                        f"This visitor comes frequently.\n"
                        f"Generate a permanent VIP code for\n"
                        f"faster entry next time?\n\n"
                        f"Click YES to create VIP code now.",
                        parent=root):
                        root.after(300, lambda n=vname, p=phone, c=company,
                                   v=visits, ph=_existing_photo[0]:
                                   _show_vip_generate_popup(n, p, c, v, ph))
        except Exception as _autofill_err:
            print(f"autofill_phone error: {_autofill_err}")
            import traceback; traceback.print_exc()
    entries["Phone"].bind("<FocusOut>", autofill_phone)

    Frame(main, bg=BORDER_CLR, height=1).pack(fill=X, side=BOTTOM)
    sbf = Frame(main, bg=STATUS_BG, pady=5)
    sbf.pack(fill=X, side=BOTTOM)
    Label(sbf, textvariable=status_var, font=("Segoe UI",9,"italic"),
          bg=STATUS_BG, fg=TEXT_MID).pack(side=LEFT, padx=14)
    tlbl = Label(sbf, text="", font=("Segoe UI",9), bg=STATUS_BG, fg=TEXT_MID)
    tlbl.pack(side=RIGHT, padx=14)
    def _tick():
        tlbl.config(text=datetime.now().strftime("🕐 %H:%M:%S  •  %d-%m-%Y"))
        root.after(1000, _tick)
    _tick()

    # ── BUTTON BAR — matches Image 1 exactly ──
    Frame(main, bg=BORDER_CLR, height=1).pack(fill=X, side=BOTTOM)
    bbar = Frame(main, bg=CARD_W, pady=12)
    bbar.pack(fill=X, side=BOTTOM)

    def _btn(text, bg, cmd):
        b = make_button(bbar, text, bg, command=cmd, padx=14, pady=11, font_size=10)
        b.pack(side=LEFT, padx=(14,0))
        return b

    _btn("  ✅  VISITOR IN  ",  BTN_IN,     lambda: save())
    out_btn = _btn("  🚪  VISITOR OUT  ", BTN_OUT, lambda: visitor_out())
    _btn("  🔍  SEARCH  ",      BTN_SEARCH, lambda: search())
    _btn("  📊  DAILY REPOS  ", BTN_REPORT, lambda: daily_report())
    _btn("  ✏  CLEAR  ",        BTN_CLEAR,  lambda: clear_fields())

    # ══════════════════════════════════════
    #  AUTOFILL PHONE
    # ══════════════════════════════════════

    # ══════════════════════════════════════
    #  AUTOFILL ID FOR OUT
    # ══════════════════════════════════════
    id_is_active_visitor = [False]

    def autofill_id_for_out(event):
        id_no = id_var.get().strip()
        id_is_active_visitor[0] = False
        if not id_no or "," in id_no: return
        try:
            df=pd.read_excel(VISITORS_FILE(),dtype=str)
            today=datetime.now().strftime(DATE_FORMAT)
            df["Date"]=df["Date"].astype(str).str.strip()
            df["Out"] =df["Out"].astype(str).str.strip()
            def iir(cell,t):
                return t in [x.strip() for x in str(cell).replace(";",",").split(",")]
            today_rows=df[df["Date"]==today]
            matches=today_rows[today_rows["ID Cards"].apply(lambda x:iir(x,id_no))]
            if matches.empty: return
            row2=matches.iloc[-1]
            out_val=str(row2.get("Out","")).strip()
            rem_str=str(row2.get("Remarks","")).strip()
            gid=str(row2.get("Group ID","")).strip()
            rem=str(row2.get("Remaining","")).strip()
            tot=str(row2.get("Total Members","")).strip()
            exited=[]
            for part in rem_str.split("|"):
                if part.strip().startswith("EXITED:"):
                    exited=[x.strip() for x in part.strip().replace("EXITED:","").split(",") if x.strip()]; break
            if out_val not in OUT_EMPTY:
                status_var.set(f"♻️  ID {id_no} — group {gid} fully exited."); return
            if id_no in exited:
                status_var.set(f"♻️  ID {id_no} — already individually exited."); return
            id_is_active_visitor[0]=True
            for col,key in [("Visitor","Visitor Name"),("Phone","Phone"),
                             ("Company","Company"),("GST No","GST No"),
                             ("Officer","Officer"),("Division","Division"),("Room No","Room No")]:
                entries[key].delete(0,END)
                entries[key].insert(0,str(row2.get(col,"")).strip())
            purp=str(row2.get("Purpose","")).strip()
            if purp and purp not in ("nan","NaN","None"): entries["Purpose"].set(purp)
            for col,key in [("Block","Block"),("Floor","Floor")]:
                if key in entries:
                    v=str(row2.get(col,"")).strip()
                    if v and v not in ("","nan","NaN","None"): entries[key].set(v)
            entries["Remarks"].delete(0,END)
            status_var.set(f"🚪  ID {id_no} → Group {gid} | Remaining: {rem}/{tot} → Click VISITOR OUT ⬅")
            out_btn.config(bg=GOLD, fg=NAVY)
            root.after(700,  lambda: out_btn.config(bg=BTN_OUT, fg=TEXT_WHITE))
            root.after(1400, lambda: out_btn.config(bg=GOLD,    fg=NAVY))
            root.after(2100, lambda: out_btn.config(bg=BTN_OUT, fg=TEXT_WHITE))
        except Exception as ex:
            status_var.set(f"Error: {ex}")
    id_entry.bind("<FocusOut>", autofill_id_for_out)

    # ══════════════════════════════════════
    #  CLEAR
    # ══════════════════════════════════════
    def clear_fields():
        for key,e in entries.items():
            try: e.delete(0,END)
            except:
                try: e.set("")
                except: pass
        id_var.set("")
        persons_var.set("0")
        id_list_lbl.config(text="IDs will appear here", fg=TEXT_LIGHT, bg="#F0F4FF")
        status_var.set("🗑  Form cleared — ready for next visitor")
        _is_returning[0] = False
        _existing_photo[0] = ""
        # Clear VIP field too
        vip_entry_var.set("")
        vip_status_lbl.config(text="")
        hide_banner()

    # ══════════════════════════════════════
    #  MAKE GROUP ID
    # ══════════════════════════════════════
    def make_group_id():
        today=datetime.now().strftime("%Y%m%d"); prefix=f"G{today}-"; nxt=1
        try:
            df=pd.read_excel(VISITORS_FILE(),dtype=str)
            existing=df["Group ID"].astype(str).str.strip()
            nums=[int(g[len(prefix):]) for g in existing
                  if g.startswith(prefix) and g[len(prefix):].isdigit()]
            if nums: nxt=max(nums)+1
        except: nxt=1
        return f"{prefix}{nxt:03d}"

    # ══════════════════════════════════════
    #  VISITOR IN
    # ══════════════════════════════════════
    def save():
        phone=entries["Phone"].get().strip()
        name=entries["Visitor Name"].get().strip()
        company=entries["Company"].get().strip()
        ids=parse_ids()
        if not phone: messagebox.showwarning("Missing","Phone number is required."); return
        if not name:  messagebox.showwarning("Missing","Visitor Name is required."); return
        if not ids:
            messagebox.showwarning("No IDs",
                "Enter at least one ID card number.\nMultiple: A101, A102, A103"); return
        if id_is_active_visitor[0]:
            messagebox.showerror("⚠️  Wrong Button!",
                "ID card(s) belong to a visitor ALREADY INSIDE today.\n\n"
                "➡  Click  VISITOR OUT  to record their exit.")
            out_btn.config(bg=ACCENT_RED,fg=TEXT_WHITE)
            root.after(400,  lambda: out_btn.config(bg=BTN_OUT,fg=TEXT_WHITE))
            root.after(800,  lambda: out_btn.config(bg=ACCENT_RED,fg=TEXT_WHITE))
            root.after(1200, lambda: out_btn.config(bg=BTN_OUT,fg=TEXT_WHITE))
            return

        # ── Same phone same day warning ──
        try:
            df_warn = get_cached_df()   # cache — no extra read
            today_w = datetime.now().strftime(DATE_FORMAT)
            df_warn["Date"]  = df_warn["Date"].astype(str).str.strip()
            df_warn["Phone"] = df_warn["Phone"].astype(str).str.strip()
            df_warn["Out"]   = df_warn["Out"].astype(str).str.strip()
            same_today    = df_warn[(df_warn["Date"]==today_w) &
                                    (df_warn["Phone"]==phone)]
            still_inside  = same_today[same_today["Out"].isin(OUT_EMPTY)]
            already_exited= same_today[~same_today["Out"].isin(OUT_EMPTY)]

            if not still_inside.empty:
                # Visitor is currently INSIDE — strong warning
                gids = ", ".join(still_inside["Group ID"].astype(str).tolist())
                if not messagebox.askyesno(
                    "⚠️  Visitor Already Inside Today",
                    f"Phone: {phone}\nName: {name}\n\n"
                    f"This visitor is currently INSIDE:\n"
                    f"Group(s): {gids}\n\n"
                    f"Click YES only if this is a new separate visit.\n"
                    f"Click NO to cancel.",
                    parent=root):
                    return
            elif not already_exited.empty:
                # Visitor came today but already left — just inform
                gids = ", ".join(already_exited["Group ID"].astype(str).tolist())
                messagebox.showinfo(
                    "ℹ️  Same Visitor — Returning Today",
                    f"Phone: {phone}\nName: {name}\n\n"
                    f"This visitor already visited today:\n"
                    f"Group(s): {gids}\n\n"
                    f"They have exited. This is a new entry.\n"
                    f"Proceeding normally.",
                    parent=root)
        except Exception as _warn_err:
            print(f"Same day check error: {_warn_err}")
        try:
            df_chk = get_cached_df()   # cache — no extra read
            today=datetime.now().strftime(DATE_FORMAT)
            df_chk["Date"]=df_chk["Date"].astype(str).str.strip()
            df_chk["Out"] =df_chk["Out"].astype(str).str.strip()
            today_chk=df_chk[df_chk["Date"]==today]
            def id_active(cell,target):
                return target in [x.strip() for x in str(cell).replace(";",",").split(",")]
            active_ids=[]
            for eid in ids:
                hits=today_chk[today_chk["ID Cards"].apply(lambda x:id_active(x,eid))]
                if hits.empty: continue
                row_h=hits.iloc[-1]; out_h=str(row_h.get("Out","")).strip()
                rem_s=str(row_h.get("Remarks","")).strip(); ex_ids=[]
                for part in rem_s.split("|"):
                    if part.strip().startswith("EXITED:"):
                        ex_ids=[x.strip() for x in part.strip().replace("EXITED:","").split(",") if x.strip()]; break
                if out_h in OUT_EMPTY and eid not in ex_ids:
                    active_ids.append(f"{eid} → group {str(row_h.get('Group ID','')).strip()}")
            if active_ids:
                if not messagebox.askyesno("⚠️  Active ID Card Detected",
                    "The following ID card(s) are currently INSIDE:\n\n"+"\n".join(active_ids)+
                    "\n\nClick YES only if this is genuinely a new visit."): return
        except: pass

        now=datetime.now(); grp_id=make_group_id(); total=len(ids)
        ids_str=", ".join(ids); date_str=now.strftime(DATE_FORMAT); arr_str=now.strftime(TIME_FORMAT)

        # Photo logic:
        # - Returning visitor → NO new photo, save empty (banner already shows old photo)
        # - First time visitor → ask if webcam available
        photo_path = ""
        if not _is_returning[0]:
            # First visit only — ask for photo
            cam_idx = detect_webcam()
            if cam_idx >= 0:
                if messagebox.askyesno("📷 Capture Visitor Photo?",
                    f"Visitor  :  {name}\nGroup    :  {grp_id}\n\n"
                    f"📷  Webcam ready!\n\nCapture photo now?\n(Click No to skip)"):
                    photo_path = capture_visitor_photo(grp_id, name, parent=root)

        data={
            "Group ID":grp_id,"Date":date_str,"Arrival":arr_str,"Out":"",
            "Visitor":name,"Phone":phone,"ID Cards":ids_str,
            "Total Members":str(total),"Remaining":str(total),"Company":company,
            "Purpose":entries["Purpose"].get().strip(),
            "GST No":entries["GST No"].get().strip(),
            "Officer":entries["Officer"].get().strip(),
            "Division":entries["Division"].get().strip(),
            "Block":entries["Block"].get().strip() if "Block" in entries else "",
            "Floor":entries["Floor"].get().strip() if "Floor" in entries else "",
            "Room No":entries["Room No"].get().strip() if "Room No" in entries else "",
            "Remarks":entries["Remarks"].get().strip(),
            # For returning visitors: save reference to existing photo folder
            # This is the FOLDER path — display will find any jpg inside it
            # Save photo FOLDER path so returning visitor banner can find it
            "Photo": (os.path.dirname(photo_path) if photo_path and os.path.isfile(photo_path)
                      else photo_path if photo_path else "")
        }
        df=pd.read_excel(VISITORS_FILE(),dtype=str)
        df=pd.concat([df,pd.DataFrame([data])],ignore_index=True)
        df.to_excel(VISITORS_FILE(), index=False, engine='openpyxl')
        photo_note="\nPhoto    :  Saved ✅" if photo_path else "\nPhoto    :  Not captured"
        status_var.set(f"✅  Group {grp_id} IN — {total} person(s) at {arr_str}")
        messagebox.showinfo("✅ Visitor IN Recorded",
            f"Group ID   :  {grp_id}\nVisitor    :  {name}\nCompany    :  {company}\n"
            f"Members    :  {total}\nID Cards   :  {ids_str}\nArrival    :  {arr_str}{photo_note}")
        invalidate_cache(); clear_fields(); update_stats()

    # ══════════════════════════════════════
    #  VISITOR OUT (full original logic)
    # ══════════════════════════════════════
    def visitor_out():
        id_input=id_var.get().strip()
        if not id_input:
            messagebox.showwarning("Missing",
                "Type ID card number(s) of exiting person(s).\n\n"
                "One person  :  A101\nAll at once :  A101, A102, A103, A104"); return
        exit_ids=list(dict.fromkeys(x.strip() for x in id_input.replace(";",",").split(",") if x.strip()))
        if not exit_ids: messagebox.showwarning("Invalid","Could not read ID number(s)."); return
        bulk_mode=len(exit_ids)>1; out_time=datetime.now().strftime(TIME_FORMAT)
        today=datetime.now().strftime(DATE_FORMAT)
        df=pd.read_excel(VISITORS_FILE(),dtype=str)
        df["Date"]=df["Date"].astype(str).str.strip()
        df["Out"] =df["Out"].astype(str).str.strip()
        today_df=df[df["Date"]==today]
        def id_in_cell(cv,t):
            return t in [x.strip() for x in str(cv).replace(";",",").split(",")]
        not_found=[];already_full=[];already_ind=[];groups={}
        for eid in exit_ids:
            hits=today_df[today_df["ID Cards"].apply(lambda x:id_in_cell(x,eid))]
            if hits.empty: not_found.append(eid); continue
            row2=hits.iloc[-1]; didx=hits.index[-1]
            gid=str(row2.get("Group ID","")).strip()
            outval=str(row2.get("Out","")).strip()
            if outval not in OUT_EMPTY: already_full.append(f"{eid}→group {gid} out@{outval}"); continue
            rem_str=str(row2.get("Remarks","")).strip(); ex=[]
            for part in rem_str.split("|"):
                if part.strip().startswith("EXITED:"):
                    ex=[x.strip() for x in part.strip().replace("EXITED:","").split(",") if x.strip()]; break
            if eid in ex: already_ind.append(eid); continue
            if gid not in groups:
                groups[gid]={"didx":didx,"row":row2,"ex":ex,"rem_str":rem_str,"exiting":[]}
            groups[gid]["exiting"].append(eid)
        if not groups:
            parts=[]
            if not_found:    parts.append("❌ Not found today:\n   "+", ".join(not_found))
            if already_full: parts.append("ℹ️ Group already fully out:\n   "+"\n   ".join(already_full))
            if already_ind:  parts.append("ℹ️ Already individually exited:\n   "+", ".join(already_ind))
            messagebox.showwarning("Nothing to Process","\n\n".join(parts)); return
        lines=[("BULK EXIT CONFIRMATION" if bulk_mode else "EXIT CONFIRMATION"),"─"*34]
        for gid,g in groups.items():
            r2=g["row"]
            all_ids=[x.strip() for x in str(r2.get("ID Cards","")).replace(";",",").split(",")]
            after=g["ex"]+g["exiting"]; remain=[x for x in all_ids if x not in after]
            lines+=[f"Group     :  {gid}",f"Visitor   :  {str(r2.get('Visitor','')).strip()}",
                    f"Company   :  {str(r2.get('Company','')).strip()}",
                    f"Arrived   :  {str(r2.get('Arrival','')).strip()}",
                    f"Exiting   :  {', '.join(g['exiting'])}  ({len(g['exiting'])} of {len(all_ids)})",
                    f"Remaining :  {len(remain)}","─"*34]
        lines.append(f"Exit Time :  {out_time}")
        if not messagebox.askyesno("Confirm EXIT","\n".join(lines)): return
        result_parts=[]
        for gid,g in groups.items():
            idx=g["didx"]; r2=g["row"]
            all_ids=[x.strip() for x in str(r2.get("ID Cards","")).replace(";",",").split(",")]
            new_ex=g["ex"]+g["exiting"]; remain=[x for x in all_ids if x not in new_ex]
            old_rem=g["rem_str"]
            for part in old_rem.split("|"):
                if part.strip().startswith("EXITED:"): old_rem=old_rem.replace(part,"").strip(" |")
            new_rem=(old_rem+" | EXITED:"+",".join(new_ex)).strip(" |")
            df.loc[idx,"Remaining"]=str(len(remain)); df.loc[idx,"Remarks"]=new_rem
            if len(remain)==0: df.loc[idx,"Out"]=out_time
            name=str(r2.get("Visitor","")).strip(); arr=str(r2.get("Arrival","")).strip()
            if len(remain)==0:
                result_parts.append(f"✅ Group {gid} ({name}) — ALL {len(all_ids)} exited\n   Arrived: {arr}  |  Final OUT: {out_time}")
            else:
                result_parts.append(f"✅ Group {gid} ({name}) — {len(g['exiting'])} exited\n   Still inside ({len(remain)}): {', '.join(remain)}")
        df.to_excel(VISITORS_FILE(), index=False, engine='openpyxl')
        final="\n\n".join(result_parts)
        if not_found:   final+=f"\n\n❌ Not found: {', '.join(not_found)}"
        if already_ind: final+=f"\nℹ️ Already exited earlier: {', '.join(already_ind)}"
        total_exited=sum(len(g["exiting"]) for g in groups.values())
        status_var.set(f"🚪  {'Bulk ' if bulk_mode else ''}EXIT — {total_exited} ID(s) at {out_time}")
        messagebox.showinfo("EXIT Recorded",final)
        invalidate_cache(); clear_fields(); update_stats()

    # ══════════════════════════════════════
    #  SEARCH
    # ══════════════════════════════════════
    def search():
        phone=entries["Phone"].get().strip()
        if not phone: messagebox.showwarning("Missing","Enter phone number to search."); return
        df=pd.read_excel(VISITORS_FILE(),dtype=str)
        result=df[df["Phone"].astype(str).str.strip()==phone]
        if result.empty: messagebox.showinfo("No Record","No visitor found with that number."); return
        top=Toplevel(root); top.title(f"Records — {phone}")
        top.geometry("1100x380"); top.configure(bg=BG_PAGE); apply_modern_style()
        hdr=Frame(top,bg=BG_HEADER,pady=10); hdr.pack(fill=X)
        Label(hdr,text=f"  🔍  Visitor Records — {phone}",
              font=("Segoe UI",11,"bold"),bg=BG_HEADER,fg=TEXT_WHITE,pady=6).pack(side=LEFT)
        t=build_treeview(top,list(df.columns))
        for r in result.itertuples(): t.insert("","END",values=list(r)[1:])
        def view_sel_photo():
            sel=t.selection()
            if not sel: return
            vals=t.item(sel[0])["values"]
            view_visitor_photo(str(vals[-1]) if vals else "",
                               str(vals[4]) if len(vals)>4 else "Visitor",parent=top)
        make_button(top,"📷  View Photo",PURPLE,command=view_sel_photo,padx=12,pady=7).pack(pady=8)

    # ══════════════════════════════════════
    #  DAILY REPORT
    # ══════════════════════════════════════
    def daily_report():
        today = datetime.now().strftime(DATE_FORMAT)
        try:
            df = pd.read_excel(VISITORS_FILE(), dtype=str)
            df["Date"] = df["Date"].astype(str).str.strip()
            rpt = df[df["Date"] == today].copy()
        except Exception as ex:
            messagebox.showerror("Error", f"Could not read data:\n{ex}"); return

        if rpt.empty:
            messagebox.showinfo("Daily Report",
                f"No visitors recorded today ({today}).\n\nReport not generated.")
            return

        # ── Ask where to save ──
        from tkinter import filedialog
        default_name = f"GST_Report_{today.replace('-','_')}"
        save_path = filedialog.asksaveasfilename(
            title="Save Daily Report",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[
                ("Excel File", "*.xlsx"),
                ("All Files",  "*.*")
            ])
        if not save_path: return  # user cancelled

        try:
            # Save styled Excel
            rpt.to_excel(save_path, index=False)
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            wb = load_workbook(save_path); ws = wb.active
            hf  = PatternFill("solid", fgColor="1E3A5F")
            hft = Font(bold=True, color="FFFFFF", name="Segoe UI", size=10)
            th  = Side(style="thin", color="D1D5DB")
            bd  = Border(left=th, right=th, top=th, bottom=th)
            for c in ws[1]:
                c.fill = hf; c.font = hft
                c.alignment = Alignment(horizontal="center"); c.border = bd
            alt = PatternFill("solid", fgColor="EFF6FF")
            for row in ws.iter_rows(min_row=2):
                for c in row:
                    c.alignment = Alignment(horizontal="center"); c.border = bd
                    if c.row % 2 == 0: c.fill = alt
            for col in ws.columns:
                ml = max((len(str(c.value)) if c.value else 0) for c in col)
                ws.column_dimensions[col[0].column_letter].width = min(ml+4, 30)
            wb.save(save_path)
            status_var.set(f"📊  Report saved: {os.path.basename(save_path)}")
            if messagebox.askyesno("✅ Report Saved",
                f"Date    : {today}\n"
                f"Visitors: {len(rpt)} group(s)\n\n"
                f"Saved to:\n{save_path}\n\n"
                f"Open file now?"):
                os.startfile(save_path)
        except Exception as ex:
            messagebox.showerror("Report Error", f"Could not save report:\n{ex}")

    root.protocol("WM_DELETE_WINDOW", lambda: do_logout(root))
    root.mainloop()


    """Subtle section heading inside the form."""
    Label(parent, text=text.upper(),
          font=("Segoe UI", 7, "bold"), bg=BG_CARD, fg=ACCENT_BLUE,
          padx=16, pady=6).grid(row=row, column=0, columnspan=2, sticky=W)


# ═══════════════════════════════════════════
#  START
# ═══════════════════════════════════════════
show_login()
